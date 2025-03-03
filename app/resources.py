import json
import os
import time
import csv
import io
import falcon
import openpyxl
import bcrypt
from jinja2 import Template
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from configure import NAME_DB, USER_DB, PASSWORD_DB, HOST_DB, PORT_DB
from connection import DbHelper
from model import Project, Assembly, Package, Vulnerability, Changelog, CVE, BDU, Stats, Breadcrumb, User, \
    AssemblyCompare, OlderAssemblies
from api.query_commands.base_query import QueryError

SECRET_ADMIN_CODE = "admin"


def timeit(method):
    def timed(*args, **kw):
        ts = time.time()
        result = method(*args, **kw)
        te = time.time()
        delta = (te - ts) * 1000
        print(f'{method.__name__} выполнялся {delta:2.2f} ms')
        return result

    return timed


def timeit_all_methods(cls):
    class NewCls:
        def __init__(self, *args, **kwargs):
            self._obj = cls(*args, **kwargs)

        def __getattribute__(self, s):
            try:
                x = super().__getattribute__(s)
            except AttributeError:
                pass
            else:
                return x
            attr = self._obj.__getattribute__(s)
            if isinstance(attr, type(self.__init__)):
                return timeit(attr)
            else:
                return attr

    return NewCls


class ProjectResource:
    def __init__(self):
        self.project = Project()

    def on_get(self, req, resp):
        print("Received request:", req.params)
        export_all = req.get_param_as_bool('export_all', default=False)
        export_format = req.get_param('format', default=None)

        start = int(req.get_param('start', default=0))
        length = int(req.get_param('length', default=10))
        search_value = req.get_param('search[value]', default='')
        order_column = req.get_param('order_column', default=None)
        order_dir = req.get_param('order_dir', default=None)

        if export_format:
            body, content_type, content_disposition = self.project.export_projects_data(
                export_format, export_all, start, length, search_value, order_column, order_dir
            )
            if body:
                resp.body = body
                resp.content_type = content_type
                if content_disposition:
                    resp.append_header('Content-Disposition', content_disposition)
                resp.status = falcon.HTTP_200
            return

        total_records = self.project.get_total_count()
        filtered_records = self.project.get_filtered_count(search_value)
        projects = self.project.get_prj_paginated(start, length, search_value, order_column, order_dir)

        resp.media = {
            'draw': req.get_param('draw', default=None),
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': projects
        }
        resp.status = falcon.HTTP_200

    def on_post(self, req, resp):
        if 'AddSubmit' in req.media:
            raise falcon.HTTPMovedPermanently("/projects/add")
        if 'DeleteSubmit' in req.media:
            prj_id = req.media.prj_id
            self.prj.delete_prj(prj_id)
        if 'ViewAssm' in req.media:
            prj_id = req.media['prj_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly")
        resp.text = self.prj.get_prj()
        projects = json.loads(resp.text)
        proj = []
        for prj in projects:
            projects[prj]['ss'] = 'odd'
            if int(prj) % 2 == 0:
                projects[prj]['ss'] = 'even'
            projects[prj]['prj_id'] = prj
            proj.append(projects[prj])
        resp.status = falcon.HTTP_OK
        resp.content_type = 'text/html'
        cwd = os.getcwd()  # Get the current working directory (cwd)
        files = os.listdir(cwd)  # Get all the files in that directory
        fp = open("uro_app/template/index.html", "r")
        tempobj = Template(fp.read())
        resp.text = tempobj.render({'projects': proj})

    """def on_put_student(self, req, resp, id):
        pass

    def on_delete_student(self, req, resp, id):
        pass"""


class ProjectDeleteResource:
    pass


class ProjectAddResource:
    def on_get(self, req, resp):
        resp.status = falcon.HTTP_OK
        resp.content_type = 'text/html'
        cwd = os.getcwd()  # Get the current working directory (cwd)
        files = os.listdir(cwd)  # Get all the files in that directory
        print("Files in %r: %s" % (cwd, files))
        fp = open("uro_app/template/add_project.html", "r")
        tempobj = Template(fp.read())
        resp.text = tempobj.render()

    def on_post(self, req, resp):
        print(req.media)
        raise falcon.HTTPMovedPermanently("/projects/add")


class AssemblyResource:
    def __init__(self):
        self.assembly = Assembly()

    def on_get(self, req, resp, prj_id):
        # Проверяем, требуется ли экспорт данных
        export_all = req.get_param_as_bool('export_all', default=False)
        export_format = req.get_param('format', default=None)  # Получаем формат экспорта

        # Пагинация, поиск и сортировка по сборкам
        start = int(req.get_param('start', default=0))
        length = int(req.get_param('length', default=10))
        search_value = req.get_param('search[value]', default='')
        order_column = req.get_param('order_column', default=None)
        order_dir = req.get_param('order_dir', default=None)

        if export_format:
            if export_all:
                # Получаем все сборки без пагинации
                assemblies = self.assembly.get_all_assm(prj_id)
            else:
                # Получаем сборки с учетом пагинации и фильтров
                assemblies = self.assembly.get_assm_paginated(prj_id, start, length, search_value, order_column,
                                                              order_dir)

            # Обрабатываем экспорт в зависимости от формата
            if export_format == 'csv':
                # Export to CSV excluding 'Assembly ID'
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(['Version', 'Description', 'Date Created'])
                for assm in assemblies:
                    writer.writerow([
                        assm['assm_version'],
                        assm['assm_desc'],
                        assm['assm_date_created']
                    ])

                resp.body = output.getvalue()
                resp.content_type = 'text/csv'
                resp.append_header('Content-Disposition', f'attachment; filename="assemblies_{prj_id}.csv"')
                resp.status = falcon.HTTP_200
                return

            elif export_format == 'excel':
                # Export to Excel excluding 'Assembly ID'
                output = io.BytesIO()
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Assemblies"
                sheet.append(['Version', 'Description', 'Date Created'])
                for assm in assemblies:
                    sheet.append([
                        assm['assm_version'],
                        assm['assm_desc'],
                        assm['assm_date_created']
                    ])
                workbook.save(output)
                resp.body = output.getvalue()
                resp.content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                resp.append_header('Content-Disposition', f'attachment; filename="assemblies_{prj_id}.xlsx"')
                resp.status = falcon.HTTP_200
                return

            elif export_format == 'pdf':
                # Export to PDF excluding 'Assembly ID'
                output = io.BytesIO()
                pdf = canvas.Canvas(output, pagesize=A4)
                pdf.setTitle(f"Assemblies for Project {prj_id}")
                pdf.setFont("Helvetica", 16)
                pdf.drawString(50, 800, f"Assemblies for Project {prj_id}")
                pdf.setFont("Helvetica", 10)
                y = 780
                pdf.drawString(50, y, "Version | Description | Date Created")
                y -= 20
                for assm in assemblies:
                    pdf.drawString(50, y, f"{assm['assm_version']} | {assm['assm_desc']} | {assm['assm_date_created']}")
                    y -= 20
                    if y < 50:
                        pdf.showPage()
                        y = 800
                pdf.save()
                resp.body = output.getvalue()
                resp.content_type = 'application/pdf'
                resp.append_header('Content-Disposition', f'attachment; filename="assemblies_{prj_id}.pdf"')
                resp.status = falcon.HTTP_200
                return

            elif export_format == 'print':
                # Export to HTML for printing excluding 'Assembly ID'
                output = io.StringIO()
                output.write("<html><body>")
                output.write(f"<h1>Assemblies for Project {prj_id}</h1>")
                output.write(
                    "<table border='1'><thead><tr><th>Version</th><th>Description</th><th>Date Created</th></tr></thead><tbody>")
                for assm in assemblies:
                    output.write(
                        f"<tr><td>{assm['assm_version']}</td><td>{assm['assm_desc']}</td><td>{assm['assm_date_created']}</td></tr>")
                output.write("</tbody></table>")
                output.write("</body></html>")
                resp.body = output.getvalue()
                resp.content_type = 'text/html'
                resp.status = falcon.HTTP_200
                return

        else:
            # Обычная обработка запроса
            total_records = self.assembly.get_total_count(prj_id)
            filtered_records = self.assembly.get_filtered_count(prj_id, search_value)
            assemblies = self.assembly.get_assm_paginated(prj_id, start, length, search_value, order_column, order_dir)

            resp.media = {
                'draw': req.get_param('draw', default=None),
                'recordsTotal': total_records,
                'recordsFiltered': filtered_records,
                'data': assemblies
            }
            resp.status = falcon.HTTP_200

    def on_post(self, req, resp, prj_id):
        if 'AddSubmit' in req.media:
            raise falcon.HTTPMovedPermanently("/projects/add")
        elif 'DeleteAssm' in req.media:
            assm_id = req.media['assm_id']
            # Здесь нужно добавить логику для удаления сборки по assm_id
        elif 'ViewPkg' in req.media:
            assm_id = req.media['assm_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly/{assm_id}/package")
        elif 'ViewVulnerability' in req.media:
            assm_id = req.media['assm_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly/{assm_id}/vulnerability/fixed")
        elif 'ViewResVulnerability' in req.media:
            assm_id = req.media['assm_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly/{assm_id}/vulnerability/resolved")
        elif 'ViewCompare' in req.media:
            assm_id = req.media['assm_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly/{assm_id}/compare/no_prev/is_current")


class AddAssmResource:
    pass


class DeleteAssmResource:
    pass


class AssemblyJointResource(object):
    def __init__(self):
        self.pkg = Package()

    def on_get(self, req, resp, prj_id, assm_id):
        self.pkg.assm_id = assm_id
        self.pkg.joint = True
        packages = self.pkg.get_pkg()
        proj = []
        for prj in packages:
            packages[prj]['ss'] = 'odd'
            if int(prj) % 2 == 0:
                packages[prj]['ss'] = 'even'
            packages[prj]['pkg_vrs_id'] = prj
            proj.append(packages[prj])
        print(packages)
        resp.status = falcon.HTTP_OK
        resp.content_type = 'text/html'
        fp = open("uro_app/template/package.html", "r")
        tempobj = Template(fp.read())
        resp.text = tempobj.render({'packages': proj})


# class AssemblyCompareResource(object):
#     def __init__(self):
#         self.pkg = Package()
#
#     def on_get(self, req, resp, prj_id, assm_id, prev, current):
#         print(11111111111)
#         self.pkg.assm_id = assm_id
#         self.pkg.difference = True
#         if prev == 'is_prev':
#             self.pkg.prev = True
#         else:
#             self.pkg.prev = False
#         if current == 'is_current':
#             self.pkg.current = True
#         else:
#             self.pkg.current = False
#         print(11111111111)
#         packages = self.pkg.get_pkg()
#         proj = []
#         for prj in packages:
#             packages[prj]['ss'] = 'odd'
#             if int(prj) % 2 == 0:
#                 packages[prj]['ss'] = 'even'
#             packages[prj]['pkg_vrs_id'] = prj
#             proj.append(packages[prj])
#         print(packages)
#         resp.status = falcon.HTTP_OK
#         resp.content_type = 'text/html'
#         # fp = open("uro_app/template/package.html", "r")
#         # tempobj = Template(fp.read())
#         # resp.text = tempobj.render({'packages': proj})
#
#     def on_post(self):
#         pass


class PackageResource:
    def __init__(self):
        self.package = Package()

    def on_get(self, req, resp, prj_id, assm_id):
        export_all = req.get_param_as_bool('export_all', default=False)
        export_format = req.get_param('format', default=None)

        start = int(req.get_param('start', default=0))
        length = int(req.get_param('length', default=10))
        search_value = req.get_param('search[value]', default='')
        order_column = req.get_param('order_column', default=None)
        order_dir = req.get_param('order_dir', default=None)
        include_joint = req.get_param_as_bool('include_joint', default=False)

        if export_format:
            body, content_type, content_disposition = self.package.export_packages_data(
                export_format, export_all, assm_id, prj_id, include_joint, start, length, search_value, order_column,
                order_dir
            )
            if body:
                resp.body = body
                resp.content_type = content_type
                if content_disposition:
                    resp.append_header('Content-Disposition', content_disposition)
                resp.status = falcon.HTTP_200
            return

        total_records = self.package.get_total_count(assm_id, include_joint)
        filtered_records = self.package.get_filtered_count(assm_id, include_joint, search_value)
        packages = self.package.get_pkg_paginated(assm_id, include_joint, start, length, search_value, order_column,
                                                  order_dir)

        resp.media = {
            'draw': req.get_param('draw', default=None),
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': packages
        }
        resp.status = falcon.HTTP_200

    def on_post(self, req, resp, prj_id, assm_id):
        if 'ViewChangelog' in req.media:
            pkg_id = req.media['pkg_vrs_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly/{assm_id}/package/{pkg_id}/changelog")
        if 'ViewVulnerability' in req.media:
            pkg_id = req.media['pkg_vrs_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly/{assm_id}/package"
                                              f"/{pkg_id}/vulnerability/fixed")
        if 'ViewResVulnerability' in req.media:
            pkg_id = req.media['pkg_vrs_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly/{assm_id}/package"
                                              f"/{pkg_id}/vulnerability/resolved")
        if 'ViewCompare' in req.media:
            assm_id = req.media['assm_id']
            raise falcon.HTTPMovedPermanently(f"/projects/{prj_id}/assembly/{assm_id}/compare")


class CVEResource:
    def __init__(self):
        self.cve = CVE()

    def on_get(self, req, resp):
        export_all = req.get_param_as_bool('export_all', default=False)
        export_format = req.get_param('format', default=None)

        start = int(req.get_param('start', default=0))
        length = int(req.get_param('length', default=10))
        search_value = req.get_param('search[value]', default='')
        order_column = req.get_param('order_column', default=None)
        order_dir = req.get_param('order_dir', default=None)

        filters = {
            'urgency': self.decode_filter(req.get_param_as_int('urgency'), 'urgency'),
            'status': self.decode_filter(req.get_param_as_int('status'), 'status'),
            'severity_level': self.decode_filter(req.get_param_as_int('severity_level'), 'severity_level'),
            'date_discovered_start': req.get_param('date_discovered_start'),
            'date_discovered_end': req.get_param('date_discovered_end'),
        }

        if export_format:
            if export_all:
                cve_list = self.cve.get_all_cve()
            else:
                cve_list = self.cve.get_cve_paginated(
                    start, length, search_value, order_column, order_dir, filters
                )

            headers = ['CVE Name', 'Package', 'Status', 'Urgency', 'Severity', 'Date Discovered', 'Description']
            export_data = [dict(cve) for cve in cve_list]

            resp.body, resp.content_type, content_disposition = self.export_data(
                export_data, export_format, 'vulnerability_log', headers
            )

            if content_disposition:
                resp.append_header('Content-Disposition', content_disposition)
            resp.status = falcon.HTTP_200
        else:
            total_records = self.cve.get_total_count()
            filtered_records = self.cve.get_filtered_count(search_value, filters)
            cve_list = self.cve.get_cve_paginated(
                start, length, search_value, order_column, order_dir, filters
            )

            resp.media = {
                'draw': req.get_param('draw', default=None),
                'recordsTotal': total_records,
                'recordsFiltered': filtered_records,
                'data': cve_list
            }
            resp.status = falcon.HTTP_200

    def decode_filter(self, value, filter_type):
        if value is None:
            return []

        mapping = {
            'urgency': {
                1 << 0: 'not yet assigned',
                1 << 1: 'unimportant',
                1 << 2: 'low',
                1 << 3: 'medium',
                1 << 4: 'high',
                1 << 5: 'end-of-life',
            },
            'status': {
                1 << 0: 'resolved',
                1 << 1: 'open',
                1 << 2: 'undetermined',
            },
            'severity_level': {
                1 << 0: 'undefined',
                1 << 1: 'unknown',
                1 << 2: 'low',
                1 << 3: 'medium',
                1 << 4: 'high',
                1 << 5: 'critical',
            },
        }

        result = []
        for bitmask, name in mapping[filter_type].items():
            if value & bitmask:
                result.append(name)

        return result

    def export_data(self, data, format_type, filename, headers):
        """Экспорт данных в CSV, Excel, PDF или HTML"""

        if format_type == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)  # Заголовки колонок
            for row in data:
                writer.writerow(row.values())

            return output.getvalue(), 'text/csv', f'attachment; filename="{filename}.csv"'

        elif format_type == 'excel':
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = filename
            sheet.append(headers)  # Заголовки колонок
            for row in data:
                sheet.append(list(row.values()))

            workbook.save(output)
            return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f'attachment; filename="{filename}.xlsx"'

        elif format_type == 'pdf':
            output = io.BytesIO()
            pdf = canvas.Canvas(output, pagesize=A4)
            pdf.setTitle(filename)
            pdf.setFont("Helvetica", 16)
            pdf.drawString(100, 800, filename.capitalize())

            pdf.setFont("Helvetica", 10)
            y = 780
            pdf.drawString(50, y, " | ".join(headers))
            y -= 20

            for row in data:
                pdf.drawString(50, y, " | ".join(str(value) for value in row.values()))
                y -= 20
                if y < 50:
                    pdf.showPage()
                    y = 800

            pdf.save()
            return output.getvalue(), 'application/pdf', f'attachment; filename="{filename}.pdf"'

        elif format_type == 'print':
            output = io.StringIO()
            output.write("<html><body>")
            output.write(f"<h1>{filename.capitalize()}</h1>")
            output.write("<table border='1'><thead><tr>")

            for header in headers:
                output.write(f"<th>{header}</th>")

            output.write("</tr></thead><tbody>")

            for row in data:
                output.write("<tr>")
                for value in row.values():
                    output.write(f"<td>{value}</td>")
                output.write("</tr>")

            output.write("</tbody></table>")
            output.write("</body></html>")
            return output.getvalue(), 'text/html', None

        else:
            raise ValueError("Unsupported export format")


class BreadcrumbResource:
    def __init__(self):
        self.breadcrumb = Breadcrumb()

    def on_get(self, req, resp, resource, resource_id):
        if resource == 'projects':
            data = self.breadcrumb.get_project_name(resource_id)
        elif resource == 'assemblies':
            data = self.breadcrumb.get_assembly_date(resource_id)
        elif resource == 'packages':
            data = self.breadcrumb.get_package_info(resource_id)
        else:
            raise falcon.HTTPNotFound(description=f"Resource '{resource}' not found")

        if data:
            resp.media = {'data': data}
            resp.status = falcon.HTTP_200
        else:
            raise falcon.HTTPNotFound(description=f"Data for {resource} with ID {resource_id} not found")


class CVELinksResource:
    def __init__(self):
        self.cve = CVE()

    def on_get(self, req, resp, cve_name):
        links = self.cve.get_cve_links(cve_name)
        resp.media = links
        resp.status = falcon.HTTP_200


class BDURource:
    def __init__(self):
        self.bdu = BDU()

    def on_get(self, req, resp):
        cve_name = req.get_param('cve_name')
        if not cve_name:
            resp.media = {'error': 'cve_name parameter is required'}
            resp.status = falcon.HTTP_400
            return

        data = self.bdu.get_bdu_data_by_cve_name(cve_name)
        if data:
            resp.media = data
            resp.status = falcon.HTTP_200
        else:
            resp.media = {'error': 'No data found'}
            resp.status = falcon.HTTP_404


class PackageCVEResource:
    def __init__(self):
        self.cve = CVE()

    def on_get(self, req, resp, prj_id, assm_id, pkg_name):
        start = int(req.get_param('start', default=0))
        length = int(req.get_param('length', default=10))
        search_value = req.get_param('search[value]', default='')

        order_column = req.get_param('order_column', default=None)
        order_dir = req.get_param('order_dir', default=None)

        filters = {
            'urgency': self.decode_filter(req.get_param_as_int('urgency'), 'urgency'),
            'status': self.decode_filter(req.get_param_as_int('status'), 'status'),
            'severity_level': self.decode_filter(req.get_param_as_int('severity_level'), 'severity_level'),
            'date_discovered_start': req.get_param('date_discovered_start'),
            'date_discovered_end': req.get_param('date_discovered_end'),
        }

        total_records = self.cve.get_total_count_for_package(pkg_name)
        filtered_records = self.cve.get_filtered_count_for_package(search_value, filters, pkg_name)

        cve_list = self.cve.get_cve_paginated_for_package(start, length, search_value, order_column, order_dir, filters,
                                                          pkg_name)

        resp.media = {
            'draw': req.get_param('draw', default=None),
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': cve_list
        }
        resp.status = falcon.HTTP_200

    def export_data(self, data, format_type, filename, headers):
        """Экспорт данных в CSV, Excel, PDF или HTML"""

        if format_type == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)  # Заголовки колонок
            for row in data:
                writer.writerow(row.values())

            return output.getvalue(), 'text/csv', f'attachment; filename="{filename}.csv"'

        elif format_type == 'excel':
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = filename
            sheet.append(headers)  # Заголовки колонок
            for row in data:
                sheet.append(list(row.values()))

            workbook.save(output)
            return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f'attachment; filename="{filename}.xlsx"'

        elif format_type == 'pdf':
            output = io.BytesIO()
            pdf = canvas.Canvas(output, pagesize=A4)
            pdf.setTitle(filename)
            pdf.setFont("Helvetica", 16)
            pdf.drawString(100, 800, filename.capitalize())

            pdf.setFont("Helvetica", 10)
            y = 780
            pdf.drawString(50, y, " | ".join(headers))
            y -= 20

            for row in data:
                pdf.drawString(50, y, " | ".join(str(value) for value in row.values()))
                y -= 20
                if y < 50:
                    pdf.showPage()
                    y = 800

            pdf.save()
            return output.getvalue(), 'application/pdf', f'attachment; filename="{filename}.pdf"'

        elif format_type == 'print':
            output = io.StringIO()
            output.write("<html><body>")
            output.write(f"<h1>{filename.capitalize()}</h1>")
            output.write("<table border='1'><thead><tr>")

            for header in headers:
                output.write(f"<th>{header}</th>")

            output.write("</tr></thead><tbody>")

            for row in data:
                output.write("<tr>")
                for value in row.values():
                    output.write(f"<td>{value}</td>")
                output.write("</tr>")

            output.write("</tbody></table>")
            output.write("</body></html>")
            return output.getvalue(), 'text/html', None

        else:
            raise ValueError("Unsupported export format")

    def decode_filter(self, value, filter_type):
        if value is None:
            return []

        mapping = {
            'urgency': {
                1 << 0: 'not yet assigned',
                1 << 1: 'unimportant',
                1 << 2: 'low',
                1 << 3: 'medium',
                1 << 4: 'high',
                1 << 5: 'end-of-life',
            },
            'status': {
                1 << 0: 'resolved',
                1 << 1: 'open',
                1 << 2: 'undetermined',
            },
            'severity_level': {
                1 << 0: 'undefined',
                1 << 1: 'unknown',
                1 << 2: 'low',
                1 << 3: 'medium',
                1 << 4: 'high',
                1 << 5: 'critical',
            },
        }

        result = []
        for bitmask, name in mapping[filter_type].items():
            if value & bitmask:
                result.append(name)

        return result


class AssemblyCveResource(object):
    def __init__(self):
        self.cve = Vulnerability()

    def on_get(self, req, resp, prj_id, assm_id, resolved):
        self.cve.assm_id = assm_id
        if resolved == 'resolved':
            self.cve.resolved = True
        resp.text = json.dumps(self.cve.get_cve())
        resp.status = falcon.HTTP_OK
        resp.content_type = falcon.MEDIA_JSON


class JointCveResource(object):
    def __init__(self):
        self.cve = Vulnerability(joint=True)

    def on_get(self, req, resp, prj_id, assm_id, resolved):
        self.cve.assm_id = assm_id
        if resolved == 'resolved':
            self.cve.resolved = True
        resp.text = self.cve.get_cve()
        resp.status = falcon.HTTP_OK
        resp.content_type = falcon.MEDIA_JSON


class ChangelogResource:
    def __init__(self):
        self.changelog = Changelog()

    def on_get(self, req, resp, prj_id, assm_id, pkg_id):
        # Проверяем, требуется ли экспорт всех данных
        export_all = req.get_param_as_bool('export_all', default=False)
        export_format = req.get_param('format', default=None)

        start = int(req.get_param('start', default=0))
        length = int(req.get_param('length', default=10))
        search_value = req.get_param('search[value]', default='')
        order_column = req.get_param('order_column', default=None)
        order_dir = req.get_param('order_dir', default=None)

        if export_format:
            if export_all:
                changelogs = self.changelog.get_all_changelog(pkg_id)
            else:
                changelogs = self.changelog.get_changelog_paginated(pkg_id, start, length, search_value, order_column,
                                                                    order_dir)

            headers = ['Version', 'Author', 'Date Added', 'Description']
            export_data = [dict(log) for log in changelogs]

            resp.body, resp.content_type, content_disposition = self.export_data(
                export_data, export_format, 'changelog', headers
            )

            if content_disposition:
                resp.append_header('Content-Disposition', content_disposition)
            resp.status = falcon.HTTP_200
        else:
            total_records = self.changelog.get_total_count(pkg_id)
            filtered_records = self.changelog.get_filtered_count(pkg_id, search_value)
            changelogs = self.changelog.get_changelog_paginated(pkg_id, start, length, search_value, order_column,
                                                                order_dir)

            resp.media = {
                'draw': req.get_param('draw', default=None),
                'recordsTotal': total_records,
                'recordsFiltered': filtered_records,
                'data': changelogs
            }
            resp.status = falcon.HTTP_200

    def export_data(self, data, format_type, filename, headers):
        """Экспорт данных в CSV, Excel, PDF или HTML"""

        if format_type == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)  # Заголовки колонок
            for row in data:
                writer.writerow(row.values())

            return output.getvalue(), 'text/csv', f'attachment; filename="{filename}.csv"'

        elif format_type == 'excel':
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = filename
            sheet.append(headers)  # Заголовки колонок
            for row in data:
                sheet.append(list(row.values()))

            workbook.save(output)
            return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f'attachment; filename="{filename}.xlsx"'

        elif format_type == 'pdf':
            output = io.BytesIO()
            pdf = canvas.Canvas(output, pagesize=A4)
            pdf.setTitle(filename)
            pdf.setFont("Helvetica", 16)
            pdf.drawString(100, 800, filename.capitalize())

            pdf.setFont("Helvetica", 10)
            y = 780
            pdf.drawString(50, y, " | ".join(headers))
            y -= 20

            for row in data:
                pdf.drawString(50, y, " | ".join(str(value) for value in row.values()))
                y -= 20
                if y < 50:
                    pdf.showPage()
                    y = 800

            pdf.save()
            return output.getvalue(), 'application/pdf', f'attachment; filename="{filename}.pdf"'

        elif format_type == 'print':
            output = io.StringIO()
            output.write("<html><body>")
            output.write(f"<h1>{filename.capitalize()}</h1>")
            output.write("<table border='1'><thead><tr>")

            for header in headers:
                output.write(f"<th>{header}</th>")

            output.write("</tr></thead><tbody>")

            for row in data:
                output.write("<tr>")
                for value in row.values():
                    output.write(f"<td>{value}</td>")
                output.write("</tr>")

            output.write("</tbody></table>")
            output.write("</body></html>")
            return output.getvalue(), 'text/html', None

        else:
            raise ValueError("Unsupported export format")


@timeit_all_methods
class ChooseAddProjectResource:
    def on_get(self, req, resp):
        resp.status = falcon.HTTP_OK
        resp.content_type = 'text/html'
        fp = open("uro_app/template/choose_add_project.html", "r")
        tempobj = Template(fp.read())
        resp.text = tempobj.render()

    def on_post(self, req, resp):
        if 'path' in req.media:
            raise falcon.HTTPMovedPermanently(f"/projects/add/path")
        else:
            raise falcon.HTTPMovedPermanently("/projects/add/input")


@timeit_all_methods
class ProjectWrongPathResource:
    def on_get(self, req, resp):
        resp.status = falcon.HTTP_OK
        resp.content_type = 'text/html'
        fp = open("uro_app/template/wrong_path.html", "r")
        tempobj = Template(fp.read())
        resp.text = tempobj.render()


@timeit_all_methods
class ProjectExisttResource:
    def on_get(self, req, input_type, resp):
        resp.status = falcon.HTTP_OK
        resp.content_type = 'text/html'
        fp = open("uro_app/template/already_exist.html", "r")
        tempobj = Template(fp.read())
        resp.text = tempobj.render()


@timeit_all_methods
class ProjectAddPathResource:
    def on_get(self, req, resp):
        resp.status = falcon.HTTP_OK
        resp.content_type = 'text/html'
        fp = open("uro_app/template/add_path_project.html", "r")
        tempobj = Template(fp.read())
        resp.text = tempobj.render()

    def on_post(self, req, resp):
        prj = Project()
        prj.l = True
        prj.path = req.media['path']
        a = prj.add_prj_path()
        if not a:
            raise falcon.HTTPMovedPermanently("/projects/wrong_path")
        else:
            raise falcon.HTTPMovedPermanently("/projects")


@timeit_all_methods
class ProjectAddInputResource:
    def on_get(self, req, resp):
        resp.status = falcon.HTTP_OK
        resp.content_type = 'text/html'
        fp = open("uro_app/template/add_project.html", "r")
        tempobj = Template(fp.read())
        resp.text = tempobj.render()

    def on_post(self, req, resp):
        prj = Project()
        prj.input = True
        prj.prj_name = req.media['project']
        prj.rel_name = req.media['release']
        prj.prj_desc = req.media['description']
        prj.vendor = req.media['vendor']
        prj.arch_name = req.media['architecture']
        a = False
        if prj.prj_name != '':
            a = prj.add_prj_input()
        else:
            print(2)
            raise falcon.HTTPMovedPermanently("/projects/add")
        if not a:
            print(1)
            raise falcon.HTTPMovedPermanently("/projects/already_exist")
        else:
            raise falcon.HTTPMovedPermanently("/projects")


class UpdateResource:
    def __init__(self):
        self.db_helper = DbHelper(NAME_DB, USER_DB, PASSWORD_DB, HOST_DB, PORT_DB)

    def on_post(self, req, resp):
        data = req.media
        resource_type = data.get('resourceType')
        row_id = data.get('id')
        column = data.get('column')
        value = data.get('value')

        # Проверяем соответствие изменяемого столбца
        if resource_type == 'project' and column in ['prj_name', 'prj_desc', 'vendor']:
            table = 'repositories.project'
            id_column = 'prj_id'
        elif resource_type == 'assembly' and column in ['assm_desc', 'assm_version']:
            table = 'repositories.assembly'
            id_column = 'assm_id'
        elif resource_type == 'package' and column in ['pkg_name', 'version', 'author_name']:
            table = 'repositories.package'
            id_column = 'pkg_vrs_id'
        else:
            resp.status = falcon.HTTP_400
            resp.media = {'message': 'Invalid resource type or column'}
            return

        # Обновление данных в базе
        update_query = f"UPDATE {table} SET {column} = %s WHERE {id_column} = %s"
        params = (value, row_id)
        self.db_helper.execute_query(update_query, params)
        self.db_helper.commit_conn()
        resp.status = falcon.HTTP_200
        resp.media = {'message': 'Data updated successfully'}


class StatsResource:
    def __init__(self):
        self.stats = Stats()

    def on_get(self, req, resp):
        data = self.stats.get_stats()
        if data:
            resp.media = data
            resp.status = falcon.HTTP_200
        else:
            resp.media = {'error': 'No stats available'}
            resp.status = falcon.HTTP_404


class RegisterResource:
    def on_post(self, req, resp):
        data = req.media
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        admin_code = data.get('admin_code', None)

        if not username or not email or not password:
            raise falcon.HTTPBadRequest(title="Ошибка", description="Не заполнены все обязательные поля")

        user_model = User()
        result = user_model.register(username, email, password, admin_code)
        if not result.get("success"):
            resp.media = result
            resp.status = falcon.HTTP_400
            return

        resp.media = result
        resp.status = falcon.HTTP_201


class LoginResource:
    def on_post(self, req, resp):
        data = req.media
        email = data.get('email')
        password = data.get('password')

        if not email or not password:
            raise falcon.HTTPBadRequest(title="Ошибка", description="Не заполнены все поля")

        user_model = User()
        user = user_model.get_by_email(email)
        if not user:
            resp.media = {"success": False, "message": "Неверный email или пароль"}
            resp.status = falcon.HTTP_400
            return

        stored_hash = user.get("password_hash")
        if not bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8')):
            resp.media = {"success": False, "message": "Неверный email или пароль"}
            resp.status = falcon.HTTP_400
            return

        resp.media = {"success": True, "user": {"username": user.get("username"), "role": user.get("role")}}
        resp.status = falcon.HTTP_200


class AssemblyCompareResource:
    def __init__(self):
        self.compare_model = AssemblyCompare()

    def on_get(self, req, resp, prj_id, assm_id, previous_assm_id):
        start = int(req.get_param('start', default=0))
        length = int(req.get_param('length', default=10))
        search_value = req.get_param('search[value]', default='')
        order_column = req.get_param('order_column', default=None)
        order_dir = req.get_param('order_dir', default=None)

        include_joint_current = req.get_param_as_bool('include_joint_current', default=False)
        include_joint_previous = req.get_param_as_bool('include_joint_previous', default=False)
        state_filter = req.get_param('compare_state_filter', default=None)

        data = self.compare_model.get_comparison_paginated(
            assm_id, previous_assm_id,
            include_joint_current, include_joint_previous,
            search_value, state_filter, order_column, order_dir, start, length
        )

        total = self.compare_model.get_total_count(assm_id, previous_assm_id,
                                                   include_joint_current, include_joint_previous)
        filtered = self.compare_model.get_filtered_count(assm_id, previous_assm_id,
                                                         include_joint_current, include_joint_previous,
                                                         search_value, state_filter)
        resp.media = {
            "draw": req.get_param('draw', default=None),
            "recordsTotal": total,
            "recordsFiltered": filtered,
            "data": data
        }
        resp.status = falcon.HTTP_200


# resources.py
class OlderAssembliesResource:
    def __init__(self):
        self.olderAssemblies = OlderAssemblies()

    def on_get(self, req, resp, prj_id, assm_id):
        data = self.olderAssemblies.get_older_assemblies(prj_id, assm_id)
        if data is not None:
            resp.media = data
            resp.status = falcon.HTTP_200
        else:
            resp.media = {'error': f'No assemblies found for project {prj_id} older than assembly {assm_id}'}
            resp.status = falcon.HTTP_404
