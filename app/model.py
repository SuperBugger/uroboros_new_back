import json
import os
import subprocess
from datetime import datetime
import io
import csv
import openpyxl
import bcrypt
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

from api.manage_commands.assembly_uploader import AssemblyUploaderApi
from api.manage_commands.project_uploader import ProjectUploaderApi
from api.query_commands.assembly_query import AssemblyApi
from api.query_commands.changelog_query import ChangelogApi
from api.query_commands.cve_query import CveApi
from api.query_commands.package_query import PackageApi
from api.query_commands.project_query import ProjectApi
from configure import NAME_DB, USER_DB, PASSWORD_DB, HOST_DB, PORT_DB, SECRET_ADMIN_CODE
from connection import DbHelper
import logging

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
for handler in logger.handlers[:]:
    logger.removeHandler(handler)
file_handler = logging.FileHandler("/tmp/report.log", mode='w', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

class Column(object):
    def __init__(self, type_column, pk=False, fk=None):
        self.type_column = type_column
        self.primary_key = pk
        self.fk_ref = fk


class Base(object):
    def __init__(self):
        self.db_helper = DbHelper(NAME_DB, USER_DB, PASSWORD_DB, HOST_DB, PORT_DB)

    def get_data(self):
        with open('data/table.txt') as f:
            return f.read()

    def command_s(self, command):
        process = subprocess.Popen(
            command,
            stdout=subprocess.PIPE
        )
        process.wait()

    def export_data(self, data, format_type, filename, headers):
        """Экспорт данных в CSV, Excel, PDF или HTML"""

        if format_type == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in data:
                writer.writerow(row.values())
            return output.getvalue(), 'text/csv', f'attachment; filename="{filename}.csv"'

        elif format_type == 'excel':
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(headers)
            for row in data:
                sheet.append(list(row.values()))
            workbook.save(output)
            return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f'attachment; filename="{filename}.xlsx"'

        elif format_type == 'pdf':
            output = io.BytesIO()
            pdf = canvas.Canvas(output, pagesize=A4)
            pdf.setTitle(filename)
            pdf.setFont("Helvetica", 12)
            y = 800
            pdf.drawString(50, y, " | ".join(headers))
            y -= 20
            for row in data:
                pdf.drawString(50, y, " | ".join(str(value) for value in row.values()))
                y -= 20
                if y < 50:
                    pdf.showPage()
                    y = 800
            pdf.save()
            return output.getvalue(), 'application/pdf', f'attachment; filename="{filename}.pdf"'

        elif format_type == 'print':
            output = io.StringIO()
            output.write("<html><body>")
            output.write(f"<h1>{filename.capitalize()}</h1>")
            output.write("<table border='1'><thead><tr>")
            for header in headers:
                output.write(f"<th>{header}</th>")
            output.write("</tr></thead><tbody>")
            for row in data:
                output.write("<tr>")
                for value in row.values():
                    output.write(f"<td>{value}</td>")
                output.write("</tr>")
            output.write("</tbody></table>")
            output.write("</body></html>")
            return output.getvalue(), 'text/html', None


class Project(Base):
    def get_total_count(self):
        query = "SELECT COUNT(*) FROM repositories.project"
        result = self.db_helper.execute_query(query)
        return result[0][0]

    def check_columns_for_search_value(self, search_value):
        search_pattern = f"%{search_value}%"
        columns_to_check = [
            {
                'col_name': 'prj_name',
                'table_name': 'repositories.project',
                'column_name': 'prj_name',
                'alias': 'p'
            },
            {
                'col_name': 'prj_desc',
                'table_name': 'repositories.project',
                'column_name': 'prj_desc',
                'alias': 'p'
            },
            {
                'col_name': 'vendor',
                'table_name': 'repositories.project',
                'column_name': 'vendor',
                'alias': 'p'
            },
            {
                'col_name': 'arch_name',
                'table_name': 'repositories.architecture',
                'column_name': 'arch_name',
                'alias': 'a'
            }
        ]

        with_clauses = []
        params = []
        for idx, col in enumerate(columns_to_check):
            with_clause = f"""
            t{idx} AS (
                SELECT '{col['col_name']}' AS col_name
                FROM {col['table_name']} {col['alias']}
                WHERE {col['alias']}.{col['column_name']} ILIKE %s
                LIMIT 1
            )
            """
            with_clauses.append(with_clause)
            params.append(search_pattern)

        with_clause_sql = ',\n'.join(with_clauses)
        union_selects = '\nUNION ALL\n'.join([f"SELECT col_name FROM t{idx}" for idx in range(len(columns_to_check))])

        query = f"""
        WITH
        {with_clause_sql}
        {union_selects}
        """
        print(query, params)
        result = self.db_helper.execute_query(query, params)
        columns_with_hits = [row[0] for row in result]
        return columns_with_hits

    def get_filtered_count(self, search_value):
        params = []
        if search_value:
            columns_with_hits = self.check_columns_for_search_value(search_value)
            print("КОЛОННЫ ДЛЯ ПОИСКА", columns_with_hits)
            if columns_with_hits:
                search_conditions = []
                search_pattern = f"%{search_value}%"
                for column in columns_with_hits:
                    if column == 'prj_name':
                        search_conditions.append("p.prj_name ILIKE %s")
                    elif column == 'prj_desc':
                        search_conditions.append("p.prj_desc ILIKE %s")
                    elif column == 'vendor':
                        search_conditions.append("p.vendor ILIKE %s")
                    elif column == 'arch_name':
                        search_conditions.append("a.arch_name ILIKE %s")
                    params.append(search_pattern)
                where_clause = " OR ".join(search_conditions)
                query = f"""
                    SELECT COUNT(*)
                    FROM repositories.project p
                    LEFT JOIN repositories.architecture a ON p.arch_id = a.arch_id
                    WHERE {where_clause}
                """
            else:
                # Если нет совпадений, возвращаем 0
                return 0
        else:
            query = "SELECT COUNT(*) FROM repositories.project p LEFT JOIN repositories.architecture a ON p.arch_id = a.arch_id"

        print(query, params)
        result = self.db_helper.execute_query(query, params)
        return result[0][0]

    def get_prj_paginated(self, start, length, search_value, order_column, order_dir):
        query = """
            SELECT 
                p.prj_id, p.prj_name, p.prj_desc, p.vendor, a.arch_name
            FROM 
                repositories.project p
            LEFT JOIN 
                repositories.architecture a ON p.arch_id = a.arch_id
        """
        params = []

        if search_value:
            columns_with_hits = self.check_columns_for_search_value(search_value)
            if columns_with_hits:
                search_conditions = []
                search_pattern = f"%{search_value}%"
                for column in columns_with_hits:
                    if column == 'prj_name':
                        search_conditions.append("p.prj_name ILIKE %s")
                    elif column == 'prj_desc':
                        search_conditions.append("p.prj_desc ILIKE %s")
                    elif column == 'vendor':
                        search_conditions.append("p.vendor ILIKE %s")
                    elif column == 'arch_name':
                        search_conditions.append("a.arch_name ILIKE %s")
                    params.append(search_pattern)
                where_clause = " OR ".join(search_conditions)
                query += f" WHERE {where_clause}"
            else:
                # Если нет совпадений, возвращаем пустой список
                return []
        else:
            # Если нет поискового значения, не добавляем WHERE
            pass

        # Добавляем сортировку
        orderable_columns = {
            'prj_id': 'p.prj_id',
            'prj_name': 'p.prj_name',
            'prj_desc': 'p.prj_desc',
            'vendor': 'p.vendor',
            'arch_name': 'a.arch_name',
        }

        if order_column and order_dir:
            sql_order_column = orderable_columns.get(order_column, 'p.prj_name')
            sql_order_dir = 'ASC' if order_dir.lower() == 'asc' else 'DESC'
            query += f" ORDER BY {sql_order_column} {sql_order_dir}"
        else:
            query += " ORDER BY p.prj_name ASC"

        query += " LIMIT %s OFFSET %s"
        params.extend([length, start])
        print("ИТОГОВЫЙ ЗАПРОС", query, params)
        result = self.db_helper.execute_query(query, params)
        if result:
            projects = [
                {
                    'prj_id': row[0],
                    'prj_name': row[1],
                    'prj_desc': row[2],
                    'vendor': row[3],
                    'arch_name': row[4]
                }
                for row in result
            ]
            return projects
        else:
            return []

    def get_all_projects(self):
        query = """
            SELECT 
                p.prj_id, p.prj_name, p.prj_desc, p.vendor, a.arch_name
            FROM 
                repositories.project p
            LEFT JOIN 
                repositories.architecture a ON p.arch_id = a.arch_id
        """
        result = self.db_helper.execute_query(query)
        if result:
            projects = [
                {
                    'prj_id': row[0],
                    'prj_name': row[1],
                    'prj_desc': row[2],
                    'vendor': row[3],
                    'arch_name': row[4]
                }
                for row in result
            ]
            return projects
        else:
            return []

    def export_projects_data(self, export_format, export_all, start=0, length=10, search_value='', order_column=None,
                             order_dir=None):
        """
        Экспорт данных о проектах в заданном формате.

        :param export_format: str, формат экспорта ('csv', 'excel', 'pdf', 'print')
        :param export_all: bool, экспортировать все данные или с учетом пагинации
        :param start: int, начальная позиция для пагинации
        :param length: int, количество записей для выборки
        :param search_value: str, строка поиска
        :param order_column: str, поле для сортировки
        :param order_dir: str, направление сортировки ('asc' или 'desc')
        :return: (body, content_type, content_disposition)
        """
        if export_all:
            projects = self.get_all_projects()
        else:
            projects = self.get_prj_paginated(start, length, search_value, order_column, order_dir)

        headers = ['Project Name', 'Description', 'Vendor', 'Architecture']
        filename = "projects"

        if export_format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for prj in projects:
                writer.writerow([
                    prj['prj_name'],
                    prj['prj_desc'],
                    prj['vendor'],
                    prj['arch_name']
                ])
            return output.getvalue(), 'text/csv', f'attachment; filename="{filename}.csv"'

        elif export_format == 'excel':
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Projects"
            sheet.append(headers)
            for prj in projects:
                sheet.append([
                    prj['prj_name'],
                    prj['prj_desc'],
                    prj['vendor'],
                    prj['arch_name']
                ])
            workbook.save(output)
            return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f'attachment; filename="{filename}.xlsx"'

        elif export_format == 'pdf':
            output = io.BytesIO()
            pdf = canvas.Canvas(output, pagesize=A4)
            pdf.setTitle("Projects")
            pdf.setFont("Helvetica", 16)
            pdf.drawString(100, 800, "Projects List")
            pdf.setFont("Helvetica", 10)
            y = 780
            pdf.drawString(50, y, "Project Name    Description    Vendor    Architecture")
            y -= 20
            for prj in projects:
                pdf.drawString(50, y,
                               f"{prj['prj_name']}    {prj['prj_desc']}    {prj['vendor']}    {prj['arch_name']}")
                y -= 20
                if y < 50:
                    pdf.showPage()
                    y = 800
            pdf.save()
            return output.getvalue(), 'application/pdf', f'attachment; filename="{filename}.pdf"'

        elif export_format == 'print':
            output = io.StringIO()
            output.write("<html><body>")
            output.write("<h1>Projects List</h1>")
            output.write(
                "<table border='1'><thead><tr><th>Project Name</th><th>Description</th><th>Vendor</th><th>Architecture</th></tr></thead><tbody>")
            for prj in projects:
                output.write(
                    f"<tr><td>{prj['prj_name']}</td><td>{prj['prj_desc']}</td><td>{prj['vendor']}</td><td>{prj['arch_name']}</td></tr>")
            output.write("</tbody></table>")
            output.write("</body></html>")
            return output.getvalue(), 'text/html', None

        return None, None, None


class Assembly(Base):

    def format_date(self, date):
        if isinstance(date, datetime):
            return date.strftime('%d-%m-%Y, %H:%M')
        elif isinstance(date, str):
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%dT%H:%M:%S.%f%z')
                return date_obj.strftime('%d-%m-%Y, %H:%M')
            except ValueError:
                return date
        return None

    def check_columns_for_search_value(self, search_value):
        search_pattern = f"%{search_value}%"
        columns_to_check = [
            {
                'col_name': 'assm_desc',
                'table_name': 'repositories.assembly',
                'column_name': 'assm_desc',
                'alias': 'a'
            },
            {
                'col_name': 'assm_version',
                'table_name': 'repositories.assembly',
                'column_name': 'assm_version',
                'alias': 'a'
            },
            {
                'col_name': 'assm_date_created',
                'table_name': 'repositories.assembly',
                'column_name': 'assm_date_created::text',
                'alias': 'a'
            }
        ]

        with_clauses = []
        params = []
        for idx, col in enumerate(columns_to_check):
            with_clause = f"""
            t{idx} AS (
                SELECT '{col['col_name']}' AS col_name
                FROM {col['table_name']} {col['alias']}
                WHERE {col['column_name']} ILIKE %s AND prj_id = %s
                LIMIT 1
            )
            """
            with_clauses.append(with_clause)
            params.extend([search_pattern, self.prj_id])

        with_clause_sql = ',\n'.join(with_clauses)
        union_selects = '\nUNION ALL\n'.join([f"SELECT col_name FROM t{idx}" for idx in range(len(columns_to_check))])

        query = f"""
        WITH
        {with_clause_sql}
        {union_selects}
        """

        result = self.db_helper.execute_query(query, params)
        columns_with_hits = [row[0] for row in result]
        return columns_with_hits

    def get_total_count(self, prj_id):
        query = "SELECT COUNT(*) FROM repositories.assembly WHERE prj_id = %s"
        result = self.db_helper.execute_query(query, (prj_id,))
        return result[0][0]

    def get_filtered_count(self, prj_id, search_value):
        self.prj_id = prj_id  # Сохраняем prj_id для использования в методе check_columns_for_search_value
        params = [prj_id]
        if search_value:
            columns_with_hits = self.check_columns_for_search_value(search_value)
            print("КОЛОННЫ ДЛЯ ПОИСКА", columns_with_hits)
            if columns_with_hits:
                search_conditions = []
                search_pattern = f"%{search_value}%"
                for column in columns_with_hits:
                    if column == 'assm_desc':
                        search_conditions.append("assm_desc ILIKE %s")
                    elif column == 'assm_version':
                        search_conditions.append("assm_version ILIKE %s")
                    elif column == 'assm_date_created':
                        search_conditions.append("assm_date_created::text ILIKE %s")
                    params.append(search_pattern)
                where_clause = " AND (" + " OR ".join(search_conditions) + ")"
                query = f"""
                    SELECT COUNT(*)
                    FROM repositories.assembly
                    WHERE prj_id = %s {where_clause}
                """
            else:
                # Если нет совпадений, возвращаем 0
                return 0
        else:
            query = "SELECT COUNT(*) FROM repositories.assembly WHERE prj_id = %s"

        result = self.db_helper.execute_query(query, params)
        return result[0][0]

    def get_assm_paginated(self, prj_id, start, length, search_value, order_column, order_dir):
        self.prj_id = prj_id  # Сохраняем prj_id для использования в методе check_columns_for_search_value
        query = """
            SELECT assm_id, assm_version, assm_desc, assm_date_created
            FROM repositories.assembly
            WHERE prj_id = %s
        """
        params = [prj_id]

        if search_value:
            columns_with_hits = self.check_columns_for_search_value(search_value)
            if columns_with_hits:
                search_conditions = []
                search_pattern = f"%{search_value}%"
                for column in columns_with_hits:
                    if column == 'assm_desc':
                        search_conditions.append("assm_desc ILIKE %s")
                    elif column == 'assm_version':
                        search_conditions.append("assm_version ILIKE %s")
                    elif column == 'assm_date_created':
                        search_conditions.append("assm_date_created::text ILIKE %s")
                    params.append(search_pattern)
                query += " AND (" + " OR ".join(search_conditions) + ")"
            else:
                # Если нет совпадений, возвращаем пустой список
                return []
        else:
            # Если нет поискового значения, не добавляем дополнительных условий
            pass

        # Добавляем сортировку
        orderable_columns = {
            'assm_id': 'assm_id',
            'assm_version': 'assm_version',
            'assm_desc': 'assm_desc',
            'assm_date_created': 'assm_date_created',
        }

        if order_column and order_dir:
            sql_order_column = orderable_columns.get(order_column, 'assm_date_created')
            sql_order_dir = 'DESC' if order_dir.lower() == 'asc' else 'ASC'
            query += f" ORDER BY {sql_order_column} {sql_order_dir}"
        else:
            query += " ORDER BY assm_date_created DESC"

        query += " LIMIT %s OFFSET %s"
        params.extend([length, start])
        print("ИТОГОВЫЙ ЗАПРОС", query, params)
        result = self.db_helper.execute_query(query, params)
        if result:
            assemblies = [
                {
                    'assm_id': row[0],
                    'assm_version': row[1],
                    'assm_desc': row[2],
                    'assm_date_created': self.format_date(row[3])
                }
                for row in result
            ]
            return assemblies
        else:
            return []

    def get_all_assm(self, prj_id):
        query = """
            SELECT assm_id, assm_version, assm_desc, assm_date_created
            FROM repositories.assembly
            WHERE prj_id = %s
        """
        params = [prj_id]

        result = self.db_helper.execute_query(query, params)

        if result:
            assemblies = [
                {
                    'assm_id': row[0],
                    'assm_version': row[1],
                    'assm_desc': row[2],
                    'assm_date_created': self.format_date(row[3])
                }
                for row in result
            ]
            return assemblies
        else:
            return []


class Package(Base):
    def format_date(self, date):
        if isinstance(date, datetime):
            return date.strftime('%d %B %Y, %H:%M')
        elif isinstance(date, str):
            date_obj = datetime.strptime(date, '%Y-%m-%dT%H:%M:%S.%f%z')
            return date_obj.strftime('%d %B %Y, %H:%M')
        return None

    def check_columns_for_search_value(self, search_value):
        search_pattern = f"%{search_value}%"
        columns_to_check = [
            {
                'col_name': 'pkg_name',
                'table_name': 'repositories.package',
                'column_name': 'pkg_name',
                'alias': 'p'
            },
            {
                'col_name': 'version',
                'table_name': 'repositories.pkg_version',
                'column_name': 'version',
                'alias': 'v'
            },
            {
                'col_name': 'author_name',
                'table_name': 'repositories.pkg_version',
                'column_name': 'author_name',
                'alias': 'v'
            }
        ]

        with_clauses = []
        params = []
        for idx, col in enumerate(columns_to_check):
            with_clause = f"""
            t{idx} AS (
                SELECT '{col['col_name']}' AS col_name
                FROM {col['table_name']} {col['alias']}
                WHERE {col['alias']}.{col['column_name']} ILIKE %s
                LIMIT 1
            )
            """
            with_clauses.append(with_clause)
            params.append(search_pattern)

        with_clause_sql = ',\n'.join(with_clauses)
        union_selects = '\nUNION ALL\n'.join([f"SELECT col_name FROM t{idx}" for idx in range(len(columns_to_check))])

        query = f"""
        WITH
        {with_clause_sql}
        SELECT col_name FROM (
            {union_selects}
        ) AS subquery
        """

        print(query, params)
        result = self.db_helper.execute_query(query, params)
        columns_with_hits = [row[0] for row in result]
        print("КОЛОНКИ В КОТОРЫХ БЫЛО НАЙДЕНО ВХОЖДЕНИЕ", columns_with_hits)
        return columns_with_hits

    def get_total_count(self, assm_id, include_joint):
        query = """
        SELECT COUNT(*)
        FROM (
            SELECT pvf.pvid AS id, pvf.assm_id AS assemblyId
            FROM (
                SELECT an.assm_id,
                       pv.pkg_id,
                       pv.pkg_vrs_id as pvid,
                       ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) as rn
                FROM repositories.assembly ca
                JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                JOIN repositories.pkg_version pv on pv.pkg_vrs_id = an.pkg_vrs_id
                WHERE ca.assm_id = %s
                      AND (a.assm_id = ca.assm_id OR (%s AND a.assm_date_created < ca.assm_date_created))
            ) as pvf
            WHERE pvf.rn = 1
        ) AS subquery;
        """
        params = [assm_id, include_joint]
        result = self.db_helper.execute_query(query, params)
        return result[0][0] if result else 0

    def get_filtered_count(self, assm_id, include_joint, search_value):
        params = {'assm_id': assm_id, 'include_joint': include_joint}

        if search_value:
            # Используем метод check_columns_for_search_value
            columns_with_hits = self.check_columns_for_search_value(search_value)
            if not columns_with_hits:
                # Если совпадений нет ни в одном столбце, возвращаем 0
                return 0
            else:
                # Формируем условия поиска только по столбцам с совпадениями
                search_conditions = []
                search_pattern = f"%{search_value}%"
                params['search_pattern'] = search_pattern
                for column in columns_with_hits:
                    if column == 'pkg_name':
                        search_conditions.append("p.pkg_name ILIKE %(search_pattern)s")
                    elif column == 'version':
                        search_conditions.append("v.version ILIKE %(search_pattern)s")
                    elif column == 'author_name':
                        search_conditions.append("v.author_name ILIKE %(search_pattern)s")
                search_clause = " AND (" + " OR ".join(search_conditions) + ")"
        else:
            search_clause = ""

        query = f"""
        SELECT COUNT(*)
        FROM (
            SELECT pvf.pvid AS id, pvf.assm_id AS assemblyId
            FROM (
                SELECT an.assm_id,
                       pv.pkg_id,
                       pv.pkg_vrs_id as pvid,
                       ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) as rn
                FROM repositories.assembly ca
                JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                JOIN repositories.pkg_version pv on pv.pkg_vrs_id = an.pkg_vrs_id
                WHERE ca.assm_id = %(assm_id)s
                      AND (a.assm_id = ca.assm_id OR (%(include_joint)s::boolean AND a.assm_date_created < ca.assm_date_created))
            ) as pvf
            JOIN repositories.pkg_version AS v ON v.pkg_vrs_id = pvf.pvid
            JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
            WHERE pvf.rn = 1
            {search_clause}
        ) AS subquery;
        """

        print(query, params)
        result = self.db_helper.execute_query(query, params)
        return result[0][0] if result else 0

    def get_pkg_paginated(self, assm_id, include_joint, start, length, search_value, order_column, order_dir):
        params = {'assm_id': assm_id, 'include_joint': include_joint}

        if search_value:
            # Используем метод check_columns_for_search_value
            columns_with_hits = self.check_columns_for_search_value(search_value)
            if not columns_with_hits:
                # Если совпадений нет ни в одном столбце, возвращаем пустой список
                return []
            else:
                # Формируем условия поиска только по столбцам с совпадениями
                search_conditions = []
                search_pattern = f"%{search_value}%"
                params['search_pattern'] = search_pattern
                for column in columns_with_hits:
                    if column == 'pkg_name':
                        search_conditions.append("p.pkg_name ILIKE %(search_pattern)s")
                    elif column == 'version':
                        search_conditions.append("v.version ILIKE %(search_pattern)s")
                    elif column == 'author_name':
                        search_conditions.append("v.author_name ILIKE %(search_pattern)s")
                search_clause = " AND (" + " OR ".join(search_conditions) + ")"
        else:
            search_clause = ""

        query = f"""
        SELECT pvf.pvid AS id, pvf.assm_id AS assemblyId,
               a.assm_date_created AS assemblyTime,
               a.assm_desc AS assemblyDescription,
               p.pkg_id AS packageId,
               p.pkg_name AS package,
               v.version,
               v.pkg_date_created AS time,
               v.author_name AS maintainer       
        FROM (
            SELECT an.assm_id,
                   pv.pkg_id,
                   pv.pkg_vrs_id as pvid,
                   ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) as rn
            FROM repositories.assembly ca
            JOIN repositories.assembly a ON a.prj_id = ca.prj_id
            JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
            JOIN repositories.pkg_version pv on pv.pkg_vrs_id = an.pkg_vrs_id
            WHERE ca.assm_id = %(assm_id)s
                  AND (a.assm_id = ca.assm_id OR (%(include_joint)s::boolean AND a.assm_date_created < ca.assm_date_created))
        ) as pvf
        JOIN repositories.pkg_version AS v ON v.pkg_vrs_id = pvf.pvid
        JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
        JOIN repositories.assembly AS a ON a.assm_id = pvf.assm_id
        WHERE pvf.rn = 1
        {search_clause}
        """

        # Определение сортировки
        orderable_columns = {
            'pkg_vrs_id': 'pvf.pvid',
            'pkg_name': 'p.pkg_name',
            'version': 'v.version',
            'author_name': 'v.author_name',
            'pkg_date_created': 'v.pkg_date_created',
            'assm_date_created': 'a.assm_date_created',
            'assm_desc': 'a.assm_desc',
        }

        if order_column and order_dir:
            sql_order_column = orderable_columns.get(order_column, 'p.pkg_name')  # По умолчанию сортировка по pkg_name
            sql_order_dir = 'ASC' if order_dir.lower() == 'asc' else 'DESC'
            query += f" ORDER BY {sql_order_column} {sql_order_dir}"
        else:
            query += " ORDER BY p.pkg_name ASC"

        # Добавляем лимит и смещение
        query += " LIMIT %(length)s OFFSET %(start)s"
        params['length'] = length
        params['start'] = start

        print("ИТОГОВЫЙ ЗАПРОС", query, params)
        result = self.db_helper.execute_query(query, params)
        if result:
            packages = [
                {
                    'pkg_vrs_id': row[0],
                    'assm_id': row[1],
                    'assm_date_created': self.format_date(row[2]),
                    'assm_desc': row[3],
                    'pkg_id': row[4],
                    'pkg_name': row[5],
                    'version': row[6],
                    'pkg_date_created': self.format_date(row[7]),
                    'author_name': row[8],
                }
                for row in result
            ]
            return packages
        else:
            return []

    def get_all_pkg(self, assm_id, prj_id, include_previous):
        current_assm_date = self.get_assm_date(assm_id)

        if include_previous:
            assembly_ids = self.get_previous_assembly_ids(prj_id, current_assm_date)
        else:
            assembly_ids = [int(assm_id)]  # Убедимся, что assm_id - число

        query = """
            SELECT
                pv.pkg_vrs_id,
                p.pkg_name,
                pv.version,
                pv.author_name,
                pv.pkg_date_created
            FROM
                repositories.assm_pkg_vrs apv
            INNER JOIN
                repositories.pkg_version pv ON apv.pkg_vrs_id = pv.pkg_vrs_id
            INNER JOIN
                repositories.package p ON pv.pkg_id = p.pkg_id
            WHERE
                apv.assm_id = ANY(%s::int[])
        """
        params = (assembly_ids,)

        result = self.db_helper.execute_query(query, params)
        if result:
            packages = [
                {
                    'pkg_vrs_id': row[0],
                    'pkg_name': row[1],
                    'version': row[2],
                    'author_name': row[3],
                    'pkg_date_created': self.format_date(row[4]),
                    # 'assm_version': row[5],
                    # 'assm_desc': row[6]
                }
                for row in result
            ]
            return packages
        else:
            return []

    def get_assm_date(self, assm_id):
        query = "SELECT assm_date_created FROM repositories.assembly WHERE assm_id = %s"
        result = self.db_helper.execute_query(query, (assm_id,))
        if result:
            return result[0][0]
        else:
            return None

    def get_previous_assembly_ids(self, prj_id, current_assm_date):
        query = """
            SELECT assm_id FROM repositories.assembly
            WHERE prj_id = %s AND assm_date_created <= %s
        """
        params = [prj_id, current_assm_date]
        result = self.db_helper.execute_query(query, params)
        return [int(row[0]) for row in result] if result else []

    def export_packages_data(self, export_format, export_all, assm_id, prj_id, include_joint, start=0, length=10,
                             search_value='', order_column=None, order_dir=None):
        """
        Экспорт данных о пакетах в заданном формате.

        :param export_format: str, формат экспорта ('csv', 'excel', 'pdf', 'print')
        :param export_all: bool, экспортировать все данные или с учетом пагинации
        :param assm_id: int, ID сборки
        :param prj_id: int, ID проекта
        :param include_joint: bool, включать ли совместные пакеты
        :param start: int, начальная позиция для пагинации
        :param length: int, количество записей для выборки
        :param search_value: str, строка поиска
        :param order_column: str, поле для сортировки
        :param order_dir: str, направление сортировки ('asc' или 'desc')
        :return: (body, content_type, content_disposition)
        """
        if export_all:
            packages = self.get_all_pkg(assm_id, prj_id, include_joint)
        else:
            packages = self.get_pkg_paginated(assm_id, include_joint, start, length, search_value, order_column,
                                              order_dir)

        headers = ['Package Name', 'Version', 'Author', 'Date Created']
        filename = f"packages_{assm_id}"

        if export_format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for pkg in packages:
                writer.writerow([
                    pkg['pkg_name'],
                    pkg['version'],
                    pkg['author_name'],
                    pkg['pkg_date_created']
                ])
            return output.getvalue(), 'text/csv', f'attachment; filename="{filename}.csv"'

        elif export_format == 'excel':
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Packages"
            sheet.append(headers)
            for pkg in packages:
                sheet.append([
                    pkg['pkg_name'],
                    pkg['version'],
                    pkg['author_name'],
                    pkg['pkg_date_created']
                ])
            workbook.save(output)
            return output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f'attachment; filename="{filename}.xlsx"'

        elif export_format == 'pdf':
            output = io.BytesIO()
            pdf = canvas.Canvas(output, pagesize=A4)
            pdf.setTitle(f"Packages for Assembly {assm_id}")
            pdf.setFont("Helvetica", 16)
            pdf.drawString(50, 800, f"Packages for Assembly {assm_id}")
            pdf.setFont("Helvetica", 10)
            y = 780
            pdf.drawString(50, y, "Package Name | Version | Author | Date Created")
            y -= 20
            for pkg in packages:
                pdf.drawString(50, y,
                               f"{pkg['pkg_name']} | {pkg['version']} | {pkg['author_name']} | {pkg['pkg_date_created']}")
                y -= 20
                if y < 50:
                    pdf.showPage()
                    y = 800
            pdf.save()
            return output.getvalue(), 'application/pdf', f'attachment; filename="{filename}.pdf"'

        elif export_format == 'print':
            output = io.StringIO()
            output.write("<html><body>")
            output.write(f"<h1>Packages for Assembly {assm_id}</h1>")
            output.write(
                "<table border='1'><thead><tr><th>Package Name</th><th>Version</th><th>Author</th><th>Date Created</th></tr></thead><tbody>")
            for pkg in packages:
                output.write(
                    f"<tr><td>{pkg['pkg_name']}</td><td>{pkg['version']}</td><td>{pkg['author_name']}</td><td>{pkg['pkg_date_created']}</td></tr>")
            output.write("</tbody></table>")
            output.write("</body></html>")
            return output.getvalue(), 'text/html', None

        return None, None, None

class CVE(Base):
    def __init__(self):
        super().__init__()

    def get_cve_links(self, cve_name):
        query = """
               SELECT source_name, template
               FROM maintenance.link_templates
           """
        result = self.db_helper.execute_query(query)
        if result:
            links = []
            for row in result:
                source_name = row[0]
                template = row[1]
                url = template.replace('{cve}', cve_name)
                links.append({'source_name': source_name, 'url': url})
            return links
        else:
            return []

    def check_columns_for_search_value(self, search_value):
        search_pattern = f"%{search_value}%"
        columns_to_check = [
            {
                'col_name': 'cve_name',
                'table_name': 'debtracker.cve',
                'column_name': 'cve_name',
                'alias': 'c'
            },
            {
                'col_name': 'pkg_name',
                'table_name': 'debtracker.package',
                'column_name': 'pkg_name',
                'alias': 'p'
            }
        ]

        with_clauses = []
        params = []
        for idx, col in enumerate(columns_to_check):
            with_clause = f"""
            t{idx} AS (
                SELECT '{col['col_name']}' AS col_name
                FROM {col['table_name']} {col['alias']}
                WHERE {col['alias']}.{col['column_name']} ILIKE %s
                LIMIT 1
            )
            """
            with_clauses.append(with_clause)
            params.append(search_pattern)

        with_clause_sql = ',\n'.join(with_clauses)
        union_selects = '\nUNION ALL\n'.join([f"SELECT col_name FROM t{idx}" for idx in range(len(columns_to_check))])

        query = f"""
        WITH
        {with_clause_sql}
        SELECT col_name FROM (
            {union_selects}
        ) AS subquery
        """

        result = self.db_helper.execute_query(query, params)
        columns_with_hits = [row[0] for row in result]
        return columns_with_hits

    def get_total_count(self):
        query = """
        SELECT COUNT(*)
        FROM debtracker.cve c
        JOIN debtracker.cve_rep cr ON c.cve_id = cr.cve_id
        LEFT JOIN debtracker.urgency u ON cr.urg_id = u.urg_id
        LEFT JOIN debtracker.status s ON cr.st_id = s.st_id
        LEFT JOIN debtracker.pkg_version pv ON cr.fixed_pkg_vrs_id = pv.pkg_vrs_id
        LEFT JOIN debtracker.package p ON pv.pkg_id = p.pkg_id
        LEFT JOIN bdu.identifier i ON i.ident_name = c.cve_name
        LEFT JOIN bdu.vul_ident vi ON vi.ident_id = i.ident_id
        LEFT JOIN bdu.vulnerability v ON v.vul_id = vi.vul_id
        """
        result = self.db_helper.execute_query(query)
        return result[0][0] if result else 0

    def get_filtered_count(self, search_value, filters):
        where_clauses = []
        params = []

        if search_value:
            # Используем метод check_columns_for_search_value
            columns_with_hits = self.check_columns_for_search_value(search_value)
            if not columns_with_hits:
                # Если совпадений нет ни в одном столбце, возвращаем 0
                return 0
            else:
                # Формируем условия поиска только по столбцам с совпадениями
                search_conditions = []
                search_pattern = f"%{search_value}%"
                for column in columns_with_hits:
                    if column == 'cve_name':
                        search_conditions.append("c.cve_name ILIKE %s")
                    elif column == 'pkg_name':
                        search_conditions.append("p.pkg_name ILIKE %s")
                    params.append(search_pattern)
                where_clauses.append("(" + " OR ".join(search_conditions) + ")")

        # Добавляем фильтры
        if filters.get('urgency'):
            urgency_placeholders = ','.join(['%s'] * len(filters['urgency']))
            where_clauses.append(f"u.urg_name IN ({urgency_placeholders})")
            params.extend(filters['urgency'])

        if filters.get('status'):
            status_placeholders = ','.join(['%s'] * len(filters['status']))
            where_clauses.append(f"s.st_name IN ({status_placeholders})")
            params.extend(filters['status'])

        if filters.get('severity_level'):
            severity_placeholders = ','.join(['%s'] * len(filters['severity_level']))
            where_clauses.append(f"v.severity_level IN ({severity_placeholders})")
            params.extend(filters['severity_level'])

        if filters.get('date_discovered_start'):
            where_clauses.append("v.date_discovered >= %s")
            params.append(filters['date_discovered_start'])

        if filters.get('date_discovered_end'):
            where_clauses.append("v.date_discovered <= %s")
            params.append(filters['date_discovered_end'])

        where_clause = ' AND '.join(where_clauses)
        if where_clause:
            where_clause = 'WHERE ' + where_clause

        query = f"""
        SELECT COUNT(*)
        FROM debtracker.cve c
        JOIN debtracker.cve_rep cr ON c.cve_id = cr.cve_id
        LEFT JOIN debtracker.urgency u ON cr.urg_id = u.urg_id
        LEFT JOIN debtracker.status s ON cr.st_id = s.st_id
        LEFT JOIN debtracker.pkg_version pv ON cr.fixed_pkg_vrs_id = pv.pkg_vrs_id
        LEFT JOIN debtracker.package p ON pv.pkg_id = p.pkg_id
        LEFT JOIN bdu.identifier i ON i.ident_name = c.cve_name
        LEFT JOIN bdu.vul_ident vi ON vi.ident_id = i.ident_id
        LEFT JOIN bdu.vulnerability v ON v.vul_id = vi.vul_id
        {where_clause}
        """

        result = self.db_helper.execute_query(query, params)
        return result[0][0] if result else 0

    def get_cve_paginated(self, start, length, search_value, order_column, order_dir, filters):
        where_clauses = []
        params = []

        if search_value:
            # Используем метод check_columns_for_search_value
            columns_with_hits = self.check_columns_for_search_value(search_value)
            if not columns_with_hits:
                # Если совпадений нет ни в одном столбце, возвращаем пустой список
                return []
            else:
                # Формируем условия поиска только по столбцам с совпадениями
                search_conditions = []
                search_pattern = f"%{search_value}%"
                for column in columns_with_hits:
                    if column == 'cve_name':
                        search_conditions.append("c.cve_name ILIKE %s")
                    elif column == 'pkg_name':
                        search_conditions.append("p.pkg_name ILIKE %s")
                    params.append(search_pattern)
                where_clauses.append("(" + " OR ".join(search_conditions) + ")")

        # Добавляем фильтры
        if filters.get('urgency'):
            urgency_placeholders = ','.join(['%s'] * len(filters['urgency']))
            where_clauses.append(f"u.urg_name IN ({urgency_placeholders})")
            params.extend(filters['urgency'])

        if filters.get('status'):
            status_placeholders = ','.join(['%s'] * len(filters['status']))
            where_clauses.append(f"s.st_name IN ({status_placeholders})")
            params.extend(filters['status'])

        if filters.get('severity_level'):
            severity_placeholders = ','.join(['%s'] * len(filters['severity_level']))
            where_clauses.append(f"v.severity_level IN ({severity_placeholders})")
            params.extend(filters['severity_level'])

        if filters.get('date_discovered_start'):
            where_clauses.append("v.date_discovered >= %s")
            params.append(filters['date_discovered_start'])

        if filters.get('date_discovered_end'):
            where_clauses.append("v.date_discovered <= %s")
            params.append(filters['date_discovered_end'])

        where_clause = ' AND '.join(where_clauses)
        if where_clause:
            where_clause = 'WHERE ' + where_clause

        # Определение сортировки
        valid_order_columns = {
            'cve_name': 'c.cve_name',
            'pkg_name': 'p.pkg_name',
            'rep_name': 'r.rep_name',
            'st_name': 's.st_name',
            'urg_name': 'u.urg_name',
            'date_discovered': 'v.date_discovered'
        }
        order_column_sql = valid_order_columns.get(order_column, 'c.cve_name')
        order_dir_sql = 'ASC' if order_dir and order_dir.lower() == 'asc' else 'DESC'

        order_clause = f"ORDER BY {order_column_sql} {order_dir_sql}"

        limit_clause = "LIMIT %s OFFSET %s"
        params.extend([length, start])

        query = f"""
        SELECT c.cve_name, p.pkg_name, r.rep_name, s.st_name, u.urg_name, v.date_discovered, c.cve_desc, v.severity_level
        FROM debtracker.cve c
        JOIN debtracker.cve_rep cr ON c.cve_id = cr.cve_id
        LEFT JOIN debtracker.urgency u ON cr.urg_id = u.urg_id
        LEFT JOIN debtracker.status s ON cr.st_id = s.st_id
        LEFT JOIN debtracker.pkg_version pv ON cr.fixed_pkg_vrs_id = pv.pkg_vrs_id
        LEFT JOIN debtracker.package p ON pv.pkg_id = p.pkg_id
        LEFT JOIN debtracker.repository r ON cr.rep_id = r.rep_id
        LEFT JOIN bdu.identifier i ON i.ident_name = c.cve_name
        LEFT JOIN bdu.vul_ident vi ON vi.ident_id = i.ident_id
        LEFT JOIN bdu.vulnerability v ON v.vul_id = vi.vul_id
        {where_clause}
        {order_clause}
        {limit_clause}
        """

        result = self.db_helper.execute_query(query, params)
        if result:
            data = [
                {
                    'cve_name': row[0],
                    'pkg_name': row[1],
                    'rep_name': row[2],
                    'st_name': row[3],
                    'urg_name': row[4],
                    'date_discovered': row[5].strftime('%Y-%m-%d') if row[5] else 'Unknown',
                    'cve_desc': row[6],
                    'severity_level': row[7],
                }
                for row in result
            ]
            return data
        else:
            return []

    def get_total_count_for_package(self, pkg_name):
        query = """
        SELECT COUNT(*)
        FROM debtracker.cve c
        LEFT JOIN debtracker.cve_rep cr ON c.cve_id = cr.cve_id
        LEFT JOIN debtracker.pkg_version pv ON cr.fixed_pkg_vrs_id = pv.pkg_vrs_id
        LEFT JOIN debtracker.package p ON pv.pkg_id = p.pkg_id
        LEFT JOIN debtracker.urgency u ON cr.urg_id = u.urg_id
        LEFT JOIN debtracker.status s ON cr.st_id = s.st_id
        LEFT JOIN bdu.identifier i ON i.ident_name = c.cve_name
        LEFT JOIN bdu.vul_ident vi ON vi.ident_id = i.ident_id
        LEFT JOIN bdu.vulnerability v ON v.vul_id = vi.vul_id
        WHERE p.pkg_name = %s
        """
        result = self.db_helper.execute_query(query, [pkg_name])
        return result[0][0] if result else 0

    def get_filtered_count_for_package(self, search_value, filters, pkg_name):
        where_clauses = ["p.pkg_name = %s"]
        params = [pkg_name]

        if search_value:
            where_clauses.append("(c.cve_name ILIKE %s)")
            params.append('%' + search_value + '%')

        if filters.get('urgency'):
            urgency_placeholders = ','.join(['%s'] * len(filters['urgency']))
            where_clauses.append(f"u.urg_name IN ({urgency_placeholders})")
            params.extend(filters['urgency'])

        if filters.get('status'):
            status_placeholders = ','.join(['%s'] * len(filters['status']))
            where_clauses.append(f"s.st_name IN ({status_placeholders})")
            params.extend(filters['status'])

        if filters.get('severity_level'):
            severity_placeholders = ','.join(['%s'] * len(filters['severity_level']))
            where_clauses.append(f"v.severity_level IN ({severity_placeholders})")
            params.extend(filters['severity_level'])

        if filters.get('date_discovered_start'):
            where_clauses.append("v.date_discovered >= %s")
            params.append(filters['date_discovered_start'])

        if filters.get('date_discovered_end'):
            where_clauses.append("v.date_discovered <= %s")
            params.append(filters['date_discovered_end'])

        where_clause = ' AND '.join(where_clauses)

        query = f"""
        SELECT COUNT(*)
        FROM debtracker.cve c
        JOIN debtracker.cve_rep cr ON c.cve_id = cr.cve_id
        JOIN debtracker.pkg_version pv ON cr.fixed_pkg_vrs_id = pv.pkg_vrs_id
        JOIN debtracker.package p on p.pkg_id = pv.pkg_id 
        LEFT JOIN debtracker.urgency u ON cr.urg_id = u.urg_id
        LEFT JOIN debtracker.status s ON cr.st_id = s.st_id
        LEFT JOIN bdu.identifier i ON i.ident_name = c.cve_name
        LEFT JOIN bdu.vul_ident vi ON vi.ident_id = i.ident_id
        LEFT JOIN bdu.vulnerability v ON v.vul_id = vi.vul_id
        WHERE {where_clause}
        """
        result = self.db_helper.execute_query(query, params)
        return result[0][0] if result else 0

    def get_cve_paginated_for_package(self, start, length, search_value, order_column, order_dir, filters, pkg_name):
        where_clauses = ["p.pkg_name = %s"]
        params = [pkg_name]

        if search_value:
            where_clauses.append("(c.cve_name ILIKE %s OR p.pkg_name ILIKE %s)")
            params.append('%' + search_value + '%')

        if filters.get('urgency'):
            urgency_placeholders = ','.join(['%s'] * len(filters['urgency']))
            where_clauses.append(f"u.urg_name IN ({urgency_placeholders})")
            params.extend(filters['urgency'])

        if filters.get('status'):
            status_placeholders = ','.join(['%s'] * len(filters['status']))
            where_clauses.append(f"s.st_name IN ({status_placeholders})")
            params.extend(filters['status'])

        if filters.get('severity_level'):
            severity_placeholders = ','.join(['%s'] * len(filters['severity_level']))
            where_clauses.append(f"v.severity_level IN ({severity_placeholders})")
            params.extend(filters['severity_level'])

        if filters.get('date_discovered_start'):
            where_clauses.append("v.date_discovered >= %s")
            params.append(filters['date_discovered_start'])

        if filters.get('date_discovered_end'):
            where_clauses.append("v.date_discovered <= %s")
            params.append(filters['date_discovered_end'])

        where_clause = ' AND '.join(where_clauses)

        valid_order_columns = {
            'cve_name': 'c.cve_name',
            'st_name': 's.st_name',
            'urg_name': 'u.urg_name',
            'severity_level': 'v.severity_level',
            'date_discovered': 'v.date_discovered'
        }
        order_column_sql = valid_order_columns.get(order_column, 'c.cve_name')
        order_dir_sql = 'ASC' if order_dir and order_dir.lower() == 'asc' else 'DESC'

        order_clause = f"ORDER BY {order_column_sql} {order_dir_sql}"

        limit_clause = "LIMIT %s OFFSET %s"
        params.extend([length, start])

        query = f"""
        SELECT c.cve_name, s.st_name, u.urg_name, v.severity_level, v.date_discovered, c.cve_desc
        FROM debtracker.cve c
        JOIN debtracker.cve_rep cr ON c.cve_id = cr.cve_id
        JOIN debtracker.pkg_version pv ON cr.fixed_pkg_vrs_id = pv.pkg_vrs_id
        JOIN debtracker.package p on p.pkg_id = pv.pkg_id 
        LEFT JOIN debtracker.urgency u ON cr.urg_id = u.urg_id
        LEFT JOIN debtracker.status s ON cr.st_id = s.st_id
        LEFT JOIN bdu.identifier i ON i.ident_name = c.cve_name
        LEFT JOIN bdu.vul_ident vi ON vi.ident_id = i.ident_id
        LEFT JOIN bdu.vulnerability v ON v.vul_id = vi.vul_id
        WHERE {where_clause}
        {order_clause}
        {limit_clause}
        """

        print(query, params)
        result = self.db_helper.execute_query(query, params)
        if result:
            data = [
                {
                    'cve_name': row[0],
                    'st_name': row[1],
                    'urg_name': row[2],
                    'severity_level': row[3],
                    'date_discovered': row[4].strftime('%Y-%m-%d') if row[4] else 'Unknown',
                    'cve_desc': row[5],
                }
                for row in result
            ]
            return data
        else:
            return []

    def get_all_cve_for_package(self, pkg_name):
        query = """
        SELECT c.cve_name, s.st_name, u.urg_name, v.severity_level, v.date_discovered, c.cve_desc
        FROM debtracker.cve c
        JOIN debtracker.cve_rep cr ON c.cve_id = cr.cve_id
        JOIN debtracker.pkg_version pv ON cr.fixed_pkg_vrs_id = pv.pkg_vrs_id
        JOIN debtracker.package p on p.pkg_id = pv.pkg_id 
        LEFT JOIN debtracker.urgency u ON cr.urg_id = u.urg_id
        LEFT JOIN debtracker.status s ON cr.st_id = s.st_id
        LEFT JOIN bdu.identifier i ON i.ident_name = c.cve_name
        LEFT JOIN bdu.vul_ident vi ON vi.ident_id = i.ident_id
        LEFT JOIN bdu.vulnerability v ON v.vul_id = vi.vul_id
        WHERE p.pkg_name = %s
        ORDER BY v.date_discovered DESC
        """

        result = self.db_helper.execute_query(query, [pkg_name])
        if result:
            return [
                {
                    'cve_name': row[0],
                    'st_name': row[1],
                    'urg_name': row[2],
                    'severity_level': row[3],
                    'date_discovered': row[4].strftime('%Y-%m-%d') if row[4] else 'Unknown',
                    'cve_desc': row[5],
                }
                for row in result
            ]
        else:
            return []

    def get_all_cve(self):
        query = """
        SELECT c.cve_name, p.pkg_name, s.st_name, u.urg_name, v.severity_level, v.date_discovered, c.cve_desc
        FROM debtracker.cve c
        JOIN debtracker.cve_rep cr ON c.cve_id = cr.cve_id
        JOIN debtracker.pkg_version pv ON cr.fixed_pkg_vrs_id = pv.pkg_vrs_id
        JOIN debtracker.package p ON pv.pkg_id = p.pkg_id
        LEFT JOIN debtracker.urgency u ON cr.urg_id = u.urg_id
        LEFT JOIN debtracker.status s ON cr.st_id = s.st_id
        LEFT JOIN bdu.identifier i ON i.ident_name = c.cve_name
        LEFT JOIN bdu.vul_ident vi ON vi.ident_id = i.ident_id
        LEFT JOIN bdu.vulnerability v ON v.vul_id = vi.vul_id
        ORDER BY v.date_discovered DESC
        """

        result = self.db_helper.execute_query(query)
        if result:
            return [
                {
                    'cve_name': row[0],
                    'pkg_name': row[1],
                    'st_name': row[2],
                    'urg_name': row[3],
                    'severity_level': row[4],
                    'date_discovered': row[5].strftime('%Y-%m-%d') if row[5] else 'Unknown',
                    'cve_desc': row[6],
                }
                for row in result
            ]
        else:
            return []


class BDU(Base):
    def __init__(self):
        super().__init__()

    def get_bdu_data_by_cve_name(self, cve_name):
        query = """
            SELECT v.vul_ident, v.vul_desc, v.date_discovered, v.cvss2_vector, v.cvss2_score, v.cvss3_vector, v.cvss3_score
            FROM bdu.vulnerability v
            JOIN bdu.vul_ident vi ON v.vul_id = vi.vul_id
            JOIN bdu.identifier i ON vi.ident_id = i.ident_id
            WHERE i.ident_name = %s
        """
        params = [cve_name]

        result = self.db_helper.execute_query(query, params)
        if result:
            return {
                'vul_ident': result[0][0],
                'vul_desc': result[0][1],
                'date_discovered': result[0][2].strftime('%Y-%m-%d') if result[0][2] else 'Недоступно',
                'cvss2_vector': result[0][3] or 'Недоступно',
                'cvss2_score': result[0][4] or 'Недоступно',
                'cvss3_vector': result[0][5] or 'Недоступно',
                'cvss3_score': result[0][6] or 'Недоступно',
            }
        else:
            return None


class Breadcrumb(Base):
    def __init__(self):
        super().__init__()

    def get_project_name(self, project_id):
        query = """
            SELECT prj_name 
            FROM repositories.project 
            WHERE prj_id = %s
        """
        result = self.db_helper.execute_query(query, (project_id,))
        return result[0][0] if result else None

    def get_assembly_date(self, assembly_id):
        query = """
            SELECT assm_date_created, assm_desc
            FROM repositories.assembly 
            WHERE assm_id = %s
        """
        result = self.db_helper.execute_query(query, (assembly_id,))
        if result:
            date = result[0][0].strftime('%d-%m-%Y, %H:%M')
            version = result[0][1] or 'unknown'
            return f"{date}: {version}"
        return None

    def get_package_info(self, package_id):
        query = """
            SELECT p.pkg_name, v.version 
            FROM (
                SELECT an.assm_id,
                   pv.pkg_id,
                   pv.pkg_vrs_id as pvid,
                   ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) as rn
                FROM repositories.assembly ca
                JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                JOIN repositories.pkg_version pv on pv.pkg_vrs_id = an.pkg_vrs_id
                WHERE ca.assm_id = 31 AND (a.assm_id = ca.assm_id)
            ) as pvf
            JOIN repositories.pkg_version AS v ON v.pkg_vrs_id = pvf.pvid
            JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
            WHERE p.pkg_name = %s
        """
        result = self.db_helper.execute_query(query, (package_id,))
        if result:
            return {'pkg_name': result[0][0], 'version': result[0][1]}
        return None


class Vulnerability(Base):
    def __init__(self, joint=False):
        super().__init__()
        self.pkg_vrs_id = None
        self.pkg_vul_id = None
        self.assm_id = None
        self.resolved = False
        self.joint = joint
        self.delete = False
        self.urgency = None
        self.status = None
        self.severity = None
        self.fdate = None
        self.sdate = None

    def filters(self, filter_list):

        st = []
        print(filter_list)
        if 'undetermined' in filter_list:
            st.append("undetermined")
        elif 'open' in filter_list:
            st.append("open")
        elif 'resolved' in filter_list:
            st.append("resolved")

        urg = []
        if 'unimportant' in filter_list:
            urg.append("unimportant")
        elif 'low' in filter_list:
            urg.append("low")
        elif 'end-of-life' in filter_list:
            urg.append("end-of-life")
        elif 'medium' in filter_list:
            urg.append("medium")
        elif 'high' in filter_list:
            urg.append("high")

        sev = []
        if 'Неизвестен' in filter_list:
            sev.append("Неизвестен")
        elif 'Низкий' in filter_list:
            sev.append("Низкий")
        elif 'Средний' in filter_list:
            sev.append("Средний")
        elif 'Высокий' in filter_list:
            sev.append("Высокий")
        elif 'Критический' in filter_list:
            sev.append("Критический")

        if filter_list['current'] != '':
            self.fdate = filter_list['current']
        if filter_list['current'] != '':
            self.sdate = filter_list['prev']

        if len(urg) != 0:
            self.urgency = urg
        if len(st) != 0:
            self.status = st
        if len(sev) != 0:
            self.severity = sev
        print(self.urgency)
        print(self.status)
        print(self.severity)

    def get_pkg_cve(self, fil_list=None):
        if fil_list is not None:
            self.filters(fil_list)
        cve = CveApi(self.db_helper)
        return cve.run(self)

    def get_assm_cve(self, fil_list=None):
        if fil_list is not None:
            self.filters(fil_list)
        cve = CveApi(self.db_helper)
        return cve.run(self)

    def get_joint_assm_cve(self, fil_list=None):
        self.joint = True
        if fil_list is not None:
            self.filters(fil_list)
        cve = CveApi(self.db_helper)
        return cve.run(self)


class Changelog(Base):
    def __init__(self):
        super().__init__()

    def get_total_count(self, pkg_id):
        query = """
            SELECT COUNT(*)
            FROM repositories.changelog c
            JOIN repositories.pkg_version pv ON c.pkg_vrs_id = pv.pkg_vrs_id
            WHERE pv.pkg_id = %s
        """
        result = self.db_helper.execute_query(query, (pkg_id,))
        return result[0][0] if result else 0

    def get_filtered_count(self, pkg_id, search_value):
        query = """
            SELECT COUNT(*)
            FROM repositories.changelog c
            JOIN repositories.pkg_version pv ON c.pkg_vrs_id = pv.pkg_vrs_id
            WHERE pv.pkg_id = %s AND (
                c.log_desc ILIKE %s OR
                c.date_added::text ILIKE %s OR
                pv.version ILIKE %s OR
                pv.author_name ILIKE %s
            )
        """
        params = (pkg_id, f'%{search_value}%', f'%{search_value}%', f'%{search_value}%', f'%{search_value}%')
        result = self.db_helper.execute_query(query, params)
        return result[0][0] if result else 0

    def get_all_changelog(self, pkg_id):
        query = """
            SELECT pv.version, pv.author_name, c.date_added, c.log_desc
            FROM repositories.changelog c
            JOIN repositories.pkg_version pv ON c.pkg_vrs_id = pv.pkg_vrs_id
            WHERE pv.pkg_id = %s
            ORDER BY c.date_added DESC
        """
        result = self.db_helper.execute_query(query, (pkg_id,))

        if result:
            return [
                {
                    'version': row[0],
                    'author_name': row[1],
                    'date_added': row[2].strftime('%Y-%m-%d %H:%M:%S') if row[2] else None,
                    'log_desc': row[3]
                }
                for row in result
            ]
        else:
            return []

    def get_changelog_paginated(self, pkg_id, start, length, search_value, order_column, order_dir):
        query = """
            SELECT pv.version, pv.author_name, c.date_added, c.log_desc
            FROM repositories.changelog c
            JOIN repositories.pkg_version pv ON c.pkg_vrs_id = pv.pkg_vrs_id
            WHERE pv.pkg_id = %s
        """
        params = [pkg_id]

        if search_value:
            query += """
                AND (c.log_desc ILIKE %s OR c.date_added::text ILIKE %s OR pv.version ILIKE %s OR pv.author_name ILIKE %s)
            """
            params.extend([f'%{search_value}%'] * 4)

        orderable_columns = {
            'version': 'pv.version',
            'author_name': 'pv.author_name',
            'date_added': 'c.date_added',
            'log_desc': 'c.log_desc'
        }

        if order_column and order_dir:
            sql_order_column = orderable_columns.get(order_column, 'c.date_added')
            query += f" ORDER BY {sql_order_column} {order_dir.upper()}"
        else:
            query += " ORDER BY c.date_added DESC"

        query += " LIMIT %s OFFSET %s"
        params.extend([length, start])

        print(query, params)

        result = self.db_helper.execute_query(query, params)
        if result:
            changelogs = [
                {
                    'version': row[0],
                    'author_name': row[1],
                    'date_added': row[2].strftime('%Y-%m-%d %H:%M:%S') if row[2] else None,
                    'log_desc': row[3]
                }
                for row in result
            ]
            return changelogs
        else:
            return []


class Stats(Base):
    def get_stats(self):
        query = """
            SELECT last_update_bdu, last_update_debtracker, total_vulnerabilities
            FROM maintenance.stats
            WHERE id = 1
        """
        result = self.db_helper.execute_query(query)
        if result:
            stats = result[0]
            return {
                'last_update_bdu': stats[0].isoformat() if stats[0] else None,
                'last_update_debtracker': stats[1].isoformat() if stats[1] else None,
                'total_vulnerabilities': stats[2]
            }
        else:
            return None


class User(Base):
    def register(self, username, email, password, admin_code=None):
        # Проверка существования пользователя с таким email
        query = "SELECT id FROM auth.users WHERE email = %s"
        result = self.db_helper.execute_query(query, (email,))
        if result:
            return {"success": False, "message": "Пользователь с таким email уже существует"}

        # Хеширование пароля
        hashed = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

        # Если введён admin_code и он совпадает, выдаем роль admin, иначе role user
        role = "admin" if admin_code and admin_code == SECRET_ADMIN_CODE else "user"

        insert_query = """
            INSERT INTO auth.users (username, email, password_hash, role)
            VALUES (%s, %s, %s, %s)
        """
        self.db_helper.execute_query(insert_query, (username, email, hashed.decode('utf-8'), role))
        self.db_helper.commit_conn()
        return {"success": True, "user": {"username": username, "role": role}}

    def get_by_email(self, email):
        query = "SELECT id, username, password_hash, role FROM auth.users WHERE email = %s"
        result = self.db_helper.execute_query(query, (email,))
        if result:
            return {
                "id": result[0][0],
                "username": result[0][1],
                "password_hash": result[0][2],
                "role": result[0][3]
            }
        return None


class AssemblyCompare(Base):

    def get_comparison_paginated(self, current_assm_id, previous_assm_id, include_joint_current, include_joint_previous,
                                 search_value, state_filter, order_column, order_dir, start, length):
        params = []
        # Для CTE "curr": сначала подставляем параметры подзапроса
        params.extend([current_assm_id, include_joint_current])
        # Формируем условие поиска для "curr"
        if search_value:
            sp = f"%{search_value}%"
            search_condition_curr = " AND (p.pkg_name ILIKE %s OR v.version ILIKE %s)"
            curr_search_params = [sp, sp]
        else:
            search_condition_curr = ""
            curr_search_params = []
        params.extend(curr_search_params)

        # Для CTE "prev": сначала параметры подзапроса
        params.extend([previous_assm_id, include_joint_previous])
        # Условие поиска для "prev"
        if search_value:
            search_condition_prev = " AND (p.pkg_name ILIKE %s OR v.version ILIKE %s)"
            prev_search_params = [sp, sp]
        else:
            search_condition_prev = ""
            prev_search_params = []
        params.extend(prev_search_params)

        base_sql = f"""
               WITH curr AS (
                 SELECT p.pkg_name,
                        v.version AS current_version,
                        v.pkg_date_created AS current_time
                 FROM (
                     SELECT an.assm_id,
                            pv.pkg_id,
                            pv.pkg_vrs_id AS pvid,
                            ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) AS rn
                     FROM repositories.assembly ca
                     JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                     JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                     JOIN repositories.pkg_version pv ON pv.pkg_vrs_id = an.pkg_vrs_id
                     WHERE ca.assm_id = %s
                       AND (a.assm_id = ca.assm_id OR (%s::boolean AND a.assm_date_created < ca.assm_date_created))
                 ) AS pvf
                 JOIN repositories.pkg_version AS v ON v.pkg_vrs_id = pvf.pvid
                 JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
                 WHERE pvf.rn = 1
                 {search_condition_curr}
               ),
               prev AS (
                 SELECT p.pkg_name,
                        v.version AS previous_version,
                        v.pkg_date_created AS previous_time
                 FROM (
                     SELECT an.assm_id,
                            pv.pkg_id,
                            pv.pkg_vrs_id AS pvid,
                            ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) AS rn
                     FROM repositories.assembly ca
                     JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                     JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                     JOIN repositories.pkg_version pv ON pv.pkg_vrs_id = an.pkg_vrs_id
                     WHERE ca.assm_id = %s
                       AND (a.assm_id = ca.assm_id OR (%s::boolean AND a.assm_date_created < ca.assm_date_created))
                 ) AS pvf
                 JOIN repositories.pkg_version AS v ON v.pkg_vrs_id = pvf.pvid
                 JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
                 WHERE pvf.rn = 1
                 {search_condition_prev}
               )
               SELECT
                 COALESCE(curr.pkg_name, prev.pkg_name) AS pkg_name,
                 CASE 
                    WHEN curr.current_version IS NOT NULL AND prev.previous_version IS NULL THEN 1
                    WHEN curr.current_version IS NULL AND prev.previous_version IS NOT NULL THEN 2
                    WHEN curr.current_version IS NOT NULL AND prev.previous_version IS NOT NULL THEN
                      CASE 
                        WHEN comp.ver > 0 THEN 3
                        WHEN comp.ver < 0 THEN 4
                      ELSE 5
                    END
                 ELSE 5
                 END AS state,
                 prev.previous_version,
                 prev.previous_time,
                 curr.current_version,
                 curr.current_time
               FROM curr
               FULL OUTER JOIN prev ON curr.pkg_name = prev.pkg_name
               LEFT JOIN LATERAL (
                SELECT CASE 
                         WHEN curr.current_version IS NOT NULL AND prev.previous_version IS NOT NULL 
                         THEN apt_version_compare(curr.current_version, prev.previous_version)
                         ELSE NULL 
                       END AS ver
            ) comp ON TRUE
           """
        # Если задан фильтр по состоянию
        if state_filter:
            state_values = [s.strip() for s in state_filter.split(',') if s.strip().isdigit()]
            if state_values:
                placeholders = ','.join(['%s'] * len(state_values))
                base_sql += f" WHERE (CASE WHEN curr.current_version IS NOT NULL AND prev.previous_version IS NULL THEN 1 WHEN curr.current_version IS NULL AND prev.previous_version IS NOT NULL THEN 2 WHEN curr.current_version IS NOT NULL AND prev.previous_version IS NOT NULL THEN (CASE WHEN comp.ver > 0 THEN 3 WHEN comp.ver < 0 THEN 4 ELSE 5 END) ELSE 5 END) IN ({placeholders})"
                params.extend(state_values)

        # Добавляем ORDER BY, LIMIT и OFFSET
        orderable_columns = {
            "pkg_name": "pkg_name",
            "state": "state",
            "previous_version": "previous_version",
            "previous_time": "previous_time",
            "current_version": "current_version",
            "current_time": "current_time"
        }
        if order_column and order_dir:
            order_by = f" ORDER BY {orderable_columns.get(order_column, 'pkg_name')} {order_dir.upper()}"
        else:
            order_by = " ORDER BY pkg_name ASC"
        base_sql += order_by
        base_sql += " LIMIT %s OFFSET %s"
        params.extend([length, start])

        print("Final Query:", base_sql, params)
        result = self.db_helper.execute_query(base_sql, params)
        data = []
        if result:
            for row in result:
                data.append({
                    "pkg_name": row[0],
                    "state": row[1],
                    "previous_version": row[2],
                    "previous_time": row[3].strftime('%Y-%m-%d %H:%M:%S') if row[3] else "",
                    "current_version": row[4],
                    "current_time": row[5].strftime('%Y-%m-%d %H:%M:%S') if row[5] else ""
                })
        return data

    def get_total_count(self, current_assm_id, previous_assm_id, include_joint_current, include_joint_previous):
        sql = f"""
            WITH curr AS (
                SELECT p.pkg_name
                FROM (
                    SELECT an.assm_id,
                           pv.pkg_id,
                           pv.pkg_vrs_id AS pvid,
                           ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) AS rn
                    FROM repositories.assembly ca
                    JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                    JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                    JOIN repositories.pkg_version pv ON pv.pkg_vrs_id = an.pkg_vrs_id
                    WHERE ca.assm_id = %s
                      AND (a.assm_id = ca.assm_id OR (%s::boolean AND a.assm_date_created < ca.assm_date_created))
                ) AS pvf
                JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
                WHERE pvf.rn = 1
            ),
            prev AS (
                SELECT p.pkg_name
                FROM (
                    SELECT an.assm_id,
                           pv.pkg_id,
                           pv.pkg_vrs_id AS pvid,
                           ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) AS rn
                    FROM repositories.assembly ca
                    JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                    JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                    JOIN repositories.pkg_version pv ON pv.pkg_vrs_id = an.pkg_vrs_id
                    WHERE ca.assm_id = %s
                      AND (a.assm_id = ca.assm_id OR (%s::boolean AND a.assm_date_created < ca.assm_date_created))
                ) AS pvf
                JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
                WHERE pvf.rn = 1
            )
            SELECT COUNT(*) FROM (
                SELECT COALESCE(curr.pkg_name, prev.pkg_name) AS pkg_name
                FROM curr
                FULL OUTER JOIN prev ON curr.pkg_name = prev.pkg_name
            ) AS combined;
        """
        params = [current_assm_id, include_joint_current, previous_assm_id, include_joint_previous]
        print("Total Count Query:", sql, params)
        result = self.db_helper.execute_query(sql, params)
        return result[0][0] if result else 0

    def get_filtered_count(self, current_assm_id, previous_assm_id, include_joint_current, include_joint_previous,
                           search_value, state_filter):
        params = []
        # Для CTE "curr": сначала параметры подзапроса
        params.extend([current_assm_id, include_joint_current])
        if search_value:
            sp = f"%{search_value}%"
            search_condition_curr = " AND (p.pkg_name ILIKE %s OR v.version ILIKE %s)"
            curr_search_params = [sp, sp]
        else:
            search_condition_curr = ""
            curr_search_params = []
        params.extend(curr_search_params)

        # Для CTE "prev": сначала параметры подзапроса
        params.extend([previous_assm_id, include_joint_previous])
        if search_value:
            search_condition_prev = " AND (p.pkg_name ILIKE %s OR v.version ILIKE %s)"
            prev_search_params = [sp, sp]
        else:
            search_condition_prev = ""
            prev_search_params = []
        params.extend(prev_search_params)

        base_sql = f"""
            WITH curr AS (
              SELECT p.pkg_name,
                     v.version AS current_version,
                     v.pkg_date_created AS current_time
              FROM (
                  SELECT an.assm_id,
                         pv.pkg_id,
                         pv.pkg_vrs_id AS pvid,
                         ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) AS rn
                  FROM repositories.assembly ca
                  JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                  JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                  JOIN repositories.pkg_version pv ON pv.pkg_vrs_id = an.pkg_vrs_id
                  WHERE ca.assm_id = %s
                    AND (a.assm_id = ca.assm_id OR (%s::boolean AND a.assm_date_created < ca.assm_date_created))
              ) AS pvf
              JOIN repositories.pkg_version AS v ON v.pkg_vrs_id = pvf.pvid
              JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
              WHERE pvf.rn = 1
              {search_condition_curr}
            ),
            prev AS (
              SELECT p.pkg_name,
                     v.version AS previous_version,
                     v.pkg_date_created AS previous_time
              FROM (
                  SELECT an.assm_id,
                         pv.pkg_id,
                         pv.pkg_vrs_id AS pvid,
                         ROW_NUMBER() OVER (PARTITION BY pv.pkg_id ORDER BY pv.pkg_date_created DESC) AS rn
                  FROM repositories.assembly ca
                  JOIN repositories.assembly a ON a.prj_id = ca.prj_id
                  JOIN repositories.assm_pkg_vrs an ON an.assm_id = a.assm_id
                  JOIN repositories.pkg_version pv ON pv.pkg_vrs_id = an.pkg_vrs_id
                  WHERE ca.assm_id = %s
                    AND (a.assm_id = ca.assm_id OR (%s::boolean AND a.assm_date_created < ca.assm_date_created))
              ) AS pvf
              JOIN repositories.pkg_version AS v ON v.pkg_vrs_id = pvf.pvid
              JOIN repositories.package p ON p.pkg_id = pvf.pkg_id
              WHERE pvf.rn = 1
              {search_condition_prev}
            )
            SELECT COUNT(*) FROM (
              SELECT COALESCE(curr.pkg_name, prev.pkg_name) AS pkg_name,
                     CASE 
                        WHEN curr.current_version IS NOT NULL AND prev.previous_version IS NULL THEN 1
                        WHEN curr.current_version IS NULL AND prev.previous_version IS NOT NULL THEN 2
                        WHEN curr.current_version IS NOT NULL AND prev.previous_version IS NOT NULL THEN
                          CASE 
                            WHEN comp.ver > 0 THEN 3
                            WHEN comp.ver < 0 THEN 4
                            ELSE 5
                          END
                     ELSE 5
                     END AS state
              FROM curr
              FULL OUTER JOIN prev ON curr.pkg_name = prev.pkg_name
              LEFT JOIN LATERAL (
                SELECT CASE 
                         WHEN curr.current_version IS NOT NULL AND prev.previous_version IS NOT NULL
                         THEN apt_version_compare(curr.current_version, prev.previous_version)
                         ELSE NULL 
                       END AS ver
              ) comp ON TRUE
            ) AS sub
        """
        # Если задан фильтр по состоянию, добавляем его на внешнем уровне
        if state_filter:
            state_values = [s.strip() for s in state_filter.split(',') if s.strip().isdigit()]
            if state_values:
                placeholders = ','.join(['%s'] * len(state_values))
                base_sql += f" WHERE state IN ({placeholders})"
                params.extend(state_values)
        final_sql = base_sql

        print("Filtered Count Query:", final_sql, params)
        result = self.db_helper.execute_query(final_sql, params)
        return result[0][0] if result else 0


class OlderAssemblies(Base):
    def get_older_assemblies(self, prj_id, assm_id):
        # Получаем дату создания выбранной сборки
        query = """
            SELECT assm_date_created 
            FROM repositories.assembly 
            WHERE assm_id = %s
        """
        result = self.db_helper.execute_query(query, [assm_id])
        if not result:
            return None
        current_date = result[0][0]

        # Выбираем все сборки проекта с датой создания меньше текущей сборки
        query2 = """
            SELECT assm_id, assm_date_created, assm_desc, assm_version
            FROM repositories.assembly
            WHERE prj_id = %s AND assm_date_created < %s
            ORDER BY assm_date_created DESC
        """
        assemblies = self.db_helper.execute_query(query2, [prj_id, current_date])
        data = []
        if assemblies:
            for row in assemblies:
                data.append({
                    "assm_id": row[0],
                    "assm_date_created": row[1].isoformat() if row[1] else "",
                    "assm_desc": row[2],
                    "assm_version": row[3]
                })
        return data


class Report(Base):
    def generate_report(self, prj_id, assm_id, previous_assm_id):
        # 1. Получение метаданных сборок
        meta_query = "SELECT assm_date_created, assm_desc FROM repositories.assembly WHERE assm_id = %s"
        current_meta = self.db_helper.execute_query(meta_query, (assm_id,))
        previous_meta = self.db_helper.execute_query(meta_query, (previous_assm_id,))
        if not current_meta or not previous_meta:
            raise Exception("Не найдены данные для одной из сборок.")
        current_date, current_desc = current_meta[0]
        previous_date, previous_desc = previous_meta[0]

        # 2. Получение имени проекта (для заголовка)
        proj_query = "SELECT prj_name FROM repositories.project WHERE prj_id = %s"
        proj_res = self.db_helper.execute_query(proj_query, (prj_id,))
        project_name = proj_res[0][0] if proj_res else "Unknown Project"

        # Формирование заголовка отчёта
        header = (f"Проект: {project_name}\n"
                  f"Разность сборок: {previous_date.strftime('%Y-%m-%d %H:%M:%S')} ({previous_desc}) ⇒ "
                  f"{current_date.strftime('%Y-%m-%d %H:%M:%S')} ({current_desc})\n")
        report_lines = [header]

        # 3. Получение всех сборок проекта между датами предыдущей и текущей сборок (включительно)
        asm_query = """
            SELECT assm_id, assm_date_created 
            FROM repositories.assembly 
            WHERE prj_id = %s 
              AND assm_date_created BETWEEN %s AND %s
            ORDER BY assm_date_created ASC
        """
        assemblies = self.db_helper.execute_query(asm_query, (prj_id, previous_date, current_date))
        if not assemblies:
            raise Exception("Не найдены сборки в указанном интервале.")
        # Собираем список ID сборок
        assembly_ids = [row[0] for row in assemblies]

        # 4. Получение цепочки версий для каждого пакета, участвующего в этих сборках.
        # Добавляем также идентификатор pkg_vrs_id для последующего запроса changelog.
        chain_query = """
            SELECT p.pkg_name, v.version, v.pkg_date_created, v.author_name, v.pkg_vrs_id
            FROM repositories.assembly a
            JOIN repositories.assm_pkg_vrs an ON a.assm_id = an.assm_id
            JOIN repositories.pkg_version v ON v.pkg_vrs_id = an.pkg_vrs_id
            JOIN repositories.package p ON p.pkg_id = v.pkg_id
            WHERE a.assm_id = ANY(%s)
            ORDER BY p.pkg_name, a.assm_date_created ASC
        """
        packages_data = self.db_helper.execute_query(chain_query, (assembly_ids,))
        chain_by_pkg = {}
        for row in packages_data:
            pkg_name, version, pkg_date, author, pkg_vrs_id = row
            chain_by_pkg.setdefault(pkg_name, []).append({
                "version": version,
                "date": pkg_date,
                "author": author,
                "pkg_vrs_id": pkg_vrs_id
            })

        # 5. Обработка каждого пакета – если крайние версии различаются, формируем блок отчёта.
        for pkg_name, versions in chain_by_pkg.items():
            # Сортируем по дате (возрастание)
            versions.sort(key=lambda x: x["date"])
            first_ver = versions[0]["version"]
            last_ver = versions[-1]["version"]
            # Проверяем изменение между крайними версиями с помощью apt_version_compare
            comp_query = "SELECT apt_version_compare(%s, %s)"
            comp_res = self.db_helper.execute_query(comp_query, (first_ver, last_ver))
            if not comp_res or comp_res[0][0] == 0:
                # Если изменений нет – пропускаем пакет
                continue
            comp_val = comp_res[0][0]
            overall_status = "повышен" if comp_val < 0 else "понижен"
            summary_line = (f"{pkg_name} - {overall_status}: "
                            f"{first_ver}({versions[0]['date'].strftime('%Y-%m-%d %H:%M:%S')})⇒"
                            f"{last_ver}({versions[-1]['date'].strftime('%Y-%m-%d %H:%M:%S')})")
            report_lines.append(summary_line)

            # 6. Для каждой пары последовательных версий (идём от последней к первой) формируем строку изменения
            # и запрашиваем соответствующие записи changelog.
            # Обрабатываем в обратном порядке: от current к предыдущим
            for i in range(len(versions) - 1, 0, -1):
                curr = versions[i]
                prev = versions[i - 1]
                # Получаем сравнение версий
                comp_pair = self.db_helper.execute_query(comp_query, (prev["version"], curr["version"]))
                if not comp_pair or comp_pair[0][0] == 0:
                    continue  # Если изменений нет – пропускаем
                pair_val = comp_pair[0][0]
                # Определяем статус для этого шага; здесь условимся:
                # Если comp < 0, то считается, что версия "добавлена" (новее), иначе – "удалён"
                step_status = "Добавлен" if pair_val < 0 else "Удалён"
                ver_line = (f"  {step_status} - Версия:{curr['version']} от {curr['author']} от "
                            f"{curr['date'].strftime('%Y-%m-%d %H:%M:%S')}")
                report_lines.append(ver_line)
                # 7. Получение записей changelog для данной версии (используем pkg_vrs_id)
                pkg_vrs_id = curr.get("pkg_vrs_id")
                if pkg_vrs_id:
                    changelog_query = """
                        SELECT log_desc 
                        FROM repositories.changelog 
                        WHERE pkg_vrs_id = %s
                        ORDER BY date_added ASC
                    """
                    changelog_res = self.db_helper.execute_query(changelog_query, (pkg_vrs_id,))
                    for cl in changelog_res:
                        # Каждая запись выводится с префиксом "* "
                        report_lines.append(f"      * {cl[0]}")
        return "\n".join(report_lines)


class ReportPackageDetails(Base):
    def get_package_details(self, pkg_name, prev_time, curr_time):
        """
        Возвращает для указанного пакета историю промежуточных версий
        и агрегированные записи ченджлога, между датами prev_time и curr_time.
        """
        query = """
        WITH intermediate AS (
          SELECT 
            pv.pkg_vrs_id,
            pv.version,
            pv.author_name,
            pv.pkg_date_created,
            c.log_desc
          from repositories.pkg_version pv 
          join repositories.package p on pv.pkg_id = p.pkg_id 
          LEFT JOIN repositories.changelog c ON c.pkg_vrs_id = pv.pkg_vrs_id
          WHERE p.pkg_name = %s
            and pv.pkg_date_created > %s
			and pv.pkg_date_created <= %s
        )
        SELECT 
          pkg_vrs_id,
          version,
          author_name,
          TO_CHAR(pkg_date_created, 'YYYY-MM-DD HH24:MI:SS') AS pv_date,
          STRING_AGG(COALESCE(log_desc, ''), E'') AS changelog
        FROM intermediate
        GROUP BY pkg_vrs_id, version, author_name, pkg_date_created
        ORDER BY pv_date DESC;
        """
        params = (pkg_name, prev_time, curr_time)
        if pkg_name == 'acpi-support':
            print("ФИНАЛЬНЫЙ ЗАПРОС ЛЯ ДОП ИНФЫ ПО ПАКЕТУ: ", query, params)

        result = self.db_helper.execute_query(query, params)
        details = []
        if result:
            for row in result:
                details.append({
                    "pkg_vrs_id": row[0],
                    "version": row[1],
                    "author": row[2],
                    "assm_date": row[3],
                    "changelog": row[4] or ""
                })
        return details

